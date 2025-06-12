# -- coding: utf-8 --
import io
import xlwt
import csv
from datetime import date
from django.http import HttpResponse, FileResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color

import copy

import logging
logger = logging.getLogger(__name__)

from django.db.models import Q, Sum, Avg, F, Func, When ,OuterRef, Subquery
from datetime import datetime, timedelta

from django.db import transaction


from .models import Bridging
from student_registration.students.models import Person
from student_registration.outreach.models import OutreachChild
from student_registration.attendances.models import CLMAttendance, CLMAttendanceStudent, CLMStudentTotalAttendance, CLMStudentAbsences
from student_registration.schools.models import (
    School,
)




def is_allowed_create(programme):
    from student_registration.schools.models import CLMRound
    try:
        current = date.today()
        current_round = CLMRound.objects.all()

        if programme == 'BLN':
            current_round = current_round.get(current_round_bln=True)
            if current_round.start_date_bln < current < current_round.end_date_bln:
                return True
            return False

        if programme == 'ABLN':
            current_round = current_round.get(current_round_abln=True)
            if current_round.start_date_abln < current < current_round.end_date_abln:
                return True
            return False

        if programme == 'CBECE':
            current_round = current_round.get(current_round_cbece=True)
            if current_round.start_date_cbece < current < current_round.end_date_cbece:
                return True
            return False

        if programme == 'Inclusion':
            current_round = current_round.get(current_round_inclusion=True)
            if current_round.start_date_inclusion < current < current_round.end_date_inclusion:
                return True
            return False

        if programme == 'RS':
            current_round = current_round.get(current_round_rs=True)
            if current_round.start_date_rs < current < current_round.end_date_rs:
                return True
            return False

        if programme == 'Outreach':
            current_round = current_round.get(current_round_outreach=True)
            if current_round.start_date_outreach < current < current_round.end_date_outreach:
                return True
            return False

        if programme == 'GeneralQuestionnaire':
            return True

        if programme == 'Bridging':
            current_round = current_round.get(current_round_bridging=True)
            if current_round.start_date_bridging < current < current_round.end_date_bridging:
                return True
            return False


    except Exception as ex:
        return False


def is_allowed_edit(programme):
    from student_registration.schools.models import CLMRound

    try:
        current = date.today()
        current_round = CLMRound.objects.all()

        if programme == 'BLN':
            current_round = current_round.get(current_round_bln=True)
            if current_round.start_date_bln_edit < current < current_round.end_date_bln_edit:
                return True
            return False

        if programme == 'ABLN':
            current_round = current_round.get(current_round_abln=True)
            if current_round.start_date_abln_edit < current < current_round.end_date_abln_edit:
                return True
            return False

        if programme == 'CBECE':
            current_round = current_round.get(current_round_cbece=True)
            if current_round.start_date_cbece_edit < current < current_round.end_date_cbece_edit:
                return True
            return False

        if programme == 'Inclusion':
            current_round = current_round.get(current_round_inclusion=True)
            if current_round.start_date_inclusion_edit < current < current_round.end_date_inclusion_edit:
                return True
            return False

        if programme == 'RS':
            current_round = current_round.get(current_round_rs=True)
            if current_round.start_date_rs_edit < current < current_round.end_date_rs_edit:
                return True
            return False

        if programme == 'Outreach':
            current_round = current_round.get(current_round_outreach=True)
            if current_round.start_date_outreach_edit < current < current_round.end_date_outreach_edit:
                return True
            return False

        if programme == 'GeneralQuestionnaire':
            return True

        if programme == 'Bridging':
            current_round = current_round.get(current_round_bridging=True)
            if current_round.start_date_bridging_edit < current < current_round.end_date_bridging_edit:
                return True
            return False

        if programme == 'MSCC':
            current_round = current_round.get(current_round_bridging=True)
            if current_round.start_date_mscc_edit < current < current_round.end_date_mscc_edit:
                return True
            return False

    except Exception as ex:
        return False


def bln_build_xls_extraction(queryset_students, queryset_fc):
    buffer = io.BytesIO()

    # Personnel
    wb_student = xlwt.Workbook(encoding='utf-8', style_compression=2)

    ws = wb_student.add_sheet('Student')

    # Sheet header, first row
    row_num_student = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = [
        'enrollment_id',
        'First time registered?',
        'Partner',
        'CLM Round',
        'Governorate',
        'District',
        'Cadaster',
        'Location',
        'Center',
        'The language supported in the program',
        'Student Address',
        'Registration level',
        'first attendance date',
        'ID number',
        'unique number',
        'First name',
        'Father name',
        'Last name',
        'Mother fullname',
        'Gender',
        'Student Nationality',
        'Student Nationality Specify',
        'Birthday - day',
        'Birthday - month',
        'Birthday - year',
        'P-Code If a child lives in a tent / Brax in a random camp',
        'Does the child have any disability or special need?',
        'Education status',
        'Miss school date',
        'Internal number',
        'RIMS  case number',
        'Source of Identification',
        'Source of Identification Specify',
        'Source of Transportation',
        'What is the educational level of the mother?',
        'What is the educational level of the father?',
        'Phone number',
        'Phone number confirm',
        'phone owner',
        'Second Phone number',
        'Second Phone number confirm',
        'Second phone owner',
        'Main Caregiver',
        'main caregiver nationality',
        'other caregiver relationship',
        'main caregiver nationality Other ',
        'Caretaker first name',
        'Caretaker middle name',
        'Caretaker last name',
        'Caretaker mother name',
        'caretaker birthday year',
        'caretaker birthday month',
        'caretaker birthday day',
        'ID Type',
        'UNHCR case number',
        'UNHCR case number confirm',
        'Parent individual ID',
        'Parent individual ID confirm',
        'Child individual ID',
        'Child individual ID confirm',
        'UNHCR recorded barcode',
        'UNHCR recorded barcode confirm',
        'Parent Lebanese ID number',
        'Parent Lebanese ID number confirm',
        'Child Lebanese ID number',
        'Child Lebanese ID number confirm',
        'Parent Syrian ID number',
        'Parent Syrian ID number confirm',
        'Child Syrian ID number',
        'Child Syrian ID number confirm',
        'Parent Palestinian ID number',
        'Parent Palestinian ID number confirm',
        'Child Palestinian ID number',
        'Child Palestinian ID number confirm',
        'ID number of the Caretaker',
        'ID number of the Caretaker confirm',
        'ID number of the child',
        'ID number of the child confirm',
        'What is the family status of the child?',
        'Does the child have children?',
        'Child number of children',
        'Does the child participate in work?',
        'What is the type of work?',
        'Please specify',
        'How many hours does this child work in a day?',
        'Child weekly income',
        'Level of participation / Absence',
        'The main barriers affecting the daily attendance and performance of the child or drop out of',
        'Please specify',
        'test_done',
        'round_complete',
        'Did the child receive basic stationery?',
        'Did the child benefit from the PSS kit?',
        'Based on the overall score, what is the recommended learning path?',
        'Please specify',
        'Did the child use Akelius program',
        'cp_referral',
        'referal_wash',
        'referal_health',
        'referal_other',
        'referal_other_specify',
        'child_received_books',
        'child_received_printout',
        'child_received_internet',
        'Please enter the number phone calls',
        'Phone call Result of follow up',
        'Please enter the number of house visits',
        'House VisitvResult of follow up',
        'Please enter the number of family visit',
        'Familiy Visit Result of follow up',
        'Parent attended visits',
        'pss session attended',
        'pss session number',
        'pss session modality',
        'pss parent attended',
        'pss parent attended other',
        'Attended covid Session?',
        'Please enter the number of sessions',
        'Please indicate modality',
        'Parent who attended the parents meeting',
        'Please specify',
        'Attended followup Session',
        'Please enter the number of sessions',
        'Please indicate modality',
        'Parent who attended the parents meeting',
        'Please specify',
        'pre test attended arabic',
        'pre test modality arabic',
        'pre test arabic',
        'pre test attended english',
        'pre test modality english',
        'pre test english',
        'pre test attended psychomotor',
        'pre test modality psychomotor',
        'pre test psychomotor',
        'pre test attended artistic',
        'pre test modality artistic',
        'pre test artistic',
        'pre test attended math',
        'pre test modality math',
        'pre test math',
        'pre test attended social',
        'pre test modality social',
        'pre test social emotional',
        'post test attended arabic',
        'post test modality arabic',
        'post test arabic',
        'post test attended english',
        'post test modality english',
        'post test english',
        'post test attended psychomotor',
        'post test modality psychomotor',
        'post test psychomotor',
        'post test attended artistic',
        'post test modality artistic',
        'post test artistic',
        'post test attended math',
        'post test modality math',
        'post test math',
        'post test attended social',
        'post test modality social',
        'post test social emotional',
        'owner',
        'modified_by',
    ]

    for col_num in range(len(columns)):
        ws.write(row_num_student, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    qs = queryset_students.extra(select={

        'pre_test_attended_arabic': "pre_test->>'BLN_ASSESSMENT/attended_arabic'",
        'pre_test_modality_arabic': "pre_test->>'BLN_ASSESSMENT/modality_arabic'",
        'pre_test_arabic': "pre_test->>'BLN_ASSESSMENT/arabic'",

        'pre_test_attended_english': "pre_test->>'BLN_ASSESSMENT/attended_english'",
        'pre_test_modality_english': "pre_test->>'BLN_ASSESSMENT/modality_english'",
        'pre_test_english': "pre_test->>'BLN_ASSESSMENT/english'",

        'pre_test_attended_psychomotor': "pre_test->>'BLN_ASSESSMENT/attended_psychomotor'",
        'pre_test_modality_psychomotor': "pre_test->>'BLN_ASSESSMENT/modality_psychomotor'",
        'pre_test_psychomotor': "pre_test->>'BLN_ASSESSMENT/psychomotor'",

        'pre_test_attended_artistic': "pre_test->>'BLN_ASSESSMENT/attended_artistic'",
        'pre_test_modality_artistic': "pre_test->>'BLN_ASSESSMENT/modality_artistic'",
        'pre_test_artistic': "pre_test->>'BLN_ASSESSMENT/artistic'",

        'pre_test_attended_math': "pre_test->>'BLN_ASSESSMENT/attended_math'",
        'pre_test_modality_math': "pre_test->>'BLN_ASSESSMENT/modality_math'",
        'pre_test_math': "pre_test->>'BLN_ASSESSMENT/math'",

        'pre_test_attended_social': "pre_test->>'BLN_ASSESSMENT/attended_social'",
        'pre_test_modality_social': "pre_test->>'BLN_ASSESSMENT/modality_social'",
        'pre_test_social_emotional': "pre_test->>'BLN_ASSESSMENT/social_emotional'",

        'post_test_attended_arabic': "post_test->>'BLN_ASSESSMENT/attended_arabic'",
        'post_test_modality_arabic': "post_test->>'BLN_ASSESSMENT/modality_arabic'",
        'post_test_arabic': "post_test->>'BLN_ASSESSMENT/arabic'",

        'post_test_attended_english': "post_test->>'BLN_ASSESSMENT/attended_english'",
        'post_test_modality_english': "post_test->>'BLN_ASSESSMENT/modality_english'",
        'post_test_english': "post_test->>'BLN_ASSESSMENT/english'",

        'post_test_attended_psychomotor': "post_test->>'BLN_ASSESSMENT/attended_psychomotor'",
        'post_test_modality_psychomotor': "post_test->>'BLN_ASSESSMENT/modality_psychomotor'",
        'post_test_psychomotor': "post_test->>'BLN_ASSESSMENT/psychomotor'",

        'post_test_attended_artistic': "post_test->>'BLN_ASSESSMENT/attended_artistic'",
        'post_test_modality_artistic': "post_test->>'BLN_ASSESSMENT/modality_artistic'",
        'post_test_artistic': "post_test->>'BLN_ASSESSMENT/artistic'",

        'post_test_attended_math': "post_test->>'BLN_ASSESSMENT/attended_math'",
        'post_test_modality_math': "post_test->>'BLN_ASSESSMENT/modality_math'",
        'post_test_math': "post_test->>'BLN_ASSESSMENT/math'",

        'post_test_attended_social': "post_test->>'BLN_ASSESSMENT/attended_social'",
        'post_test_modality_social': "post_test->>'BLN_ASSESSMENT/modality_social'",
        'post_test_social_emotional': "post_test->>'BLN_ASSESSMENT/social_emotional'",
    })

    rows = qs.order_by('id').values_list(
        'id',
        'new_registry',
        'partner__name',
        'round__name',
        'governorate__name_en',
        'district__name_en',
        'cadaster__name_en',
        'location',
        'center__name',
        'language',
        'student__address',
        'registration_level',
        'first_attendance_date',
        'student__id_number',
        'student__number',
        'student__first_name',
        'student__father_name',
        'student__last_name',
        'student__mother_fullname',
        'student__sex',
        'student__nationality__name',
        'other_nationality',
        'student__birthday_day',
        'student__birthday_month',
        'student__birthday_year',
        'student__p_code',
        'disability__name_en',
        'education_status',
        'miss_school_date',
        'internal_number',
        'rims_case_number',
        'source_of_identification',
        'source_of_identification_specify',
        'source_of_transportation',
        'hh_educational_level__name',
        'father_educational_level__name',
        'phone_number',
        'phone_number_confirm',
        'phone_owner',
        'second_phone_number',
        'second_phone_number_confirm',
        'second_phone_owner',
        'main_caregiver',
        'main_caregiver_nationality__name',
        'other_caregiver_relationship',
        'main_caregiver_nationality_other',
        'caretaker_first_name',
        'caretaker_middle_name',
        'caretaker_last_name',
        'caretaker_mother_name',
        'caretaker_birthday_year',
        'caretaker_birthday_month',
        'caretaker_birthday_day',
        'id_type',
        'case_number',
        'case_number_confirm',
        'parent_individual_case_number',
        'parent_individual_case_number_confirm',
        'individual_case_number',
        'individual_case_number_confirm',
        'recorded_number',
        'recorded_number_confirm',
        'parent_national_number',
        'parent_national_number_confirm',
        'national_number',
        'national_number_confirm',
        'parent_syrian_national_number',
        'parent_syrian_national_number_confirm',
        'syrian_national_number',
        'syrian_national_number_confirm',
        'parent_sop_national_number',
        'parent_sop_national_number_confirm',
        'sop_national_number',
        'sop_national_number_confirm',
        'parent_other_number',
        'parent_other_number_confirm',
        'other_number',
        'other_number_confirm',
        'student__family_status',
        'student__have_children',
        'student_number_children',
        'have_labour_single_selection',
        'labours_single_selection',
        'labours_other_specify',
        'labour_hours',
        'labour_weekly_income',
        'participation',
        'barriers_single',
        'barriers_other',
        'test_done',
        'round_complete',
        'basic_stationery',
        'pss_kit',
        'learning_result',
        'learning_result_other',
        'akelius_program',
        'cp_referral',
        'referal_wash',
        'referal_health',
        'referal_other',
        'referal_other_specify',
        'child_received_books',
        'child_received_printout',
        'child_received_internet',
        'phone_call_number',
        'phone_call_follow_up_result',
        'house_visit_number',
        'house_visit_follow_up_result',
        'family_visit_number',
        'family_visit_follow_up_result',
        'parent_attended_visits',
        'pss_session_attended',
        'pss_session_number',
        'pss_session_modality',
        'pss_parent_attended',
        'pss_parent_attended_other',
        'covid_session_attended',
        'covid_session_number',
        'covid_session_modality',
        'covid_parent_attended',
        'covid_parent_attended_other',
        'followup_session_attended',
        'followup_session_number',
        'followup_session_modality',
        'followup_parent_attended',
        'followup_parent_attended_other',
        'pre_test_attended_arabic',
        'pre_test_modality_arabic',
        'pre_test_arabic',
        'pre_test_attended_english',
        'pre_test_modality_english',
        'pre_test_english',
        'pre_test_attended_psychomotor',
        'pre_test_modality_psychomotor',
        'pre_test_psychomotor',
        'pre_test_attended_artistic',
        'pre_test_modality_artistic',
        'pre_test_artistic',
        'pre_test_attended_math',
        'pre_test_modality_math',
        'pre_test_math',
        'pre_test_attended_social',
        'pre_test_modality_social',
        'pre_test_social_emotional',
        'post_test_attended_arabic',
        'post_test_modality_arabic',
        'post_test_arabic',
        'post_test_attended_english',
        'post_test_modality_english',
        'post_test_english',
        'post_test_attended_psychomotor',
        'post_test_modality_psychomotor',
        'post_test_psychomotor',
        'post_test_attended_artistic',
        'post_test_modality_artistic',
        'post_test_artistic',
        'post_test_attended_math',
        'post_test_modality_math',
        'post_test_math',
        'post_test_attended_social',
        'post_test_modality_social',
        'post_test_social_emotional',
        'owner__username',
        'modified_by__username',
    )
    for row in rows:
        row_num_student += 1
        for col_num in range(len(row)):
            ws.write(row_num_student, col_num, row[col_num], font_style)
    # FC
    wsFC = wb_student.add_sheet('FC')
    row_num_fc = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    column_header_fc = [
        'enrollment id',
        'Partner',
        'CLM Round',
        'First time registered?',
        'Registration level',
        'Student first name',
        'Student father name',
        'Student last name',
        'unique number',
        'Gender',
        'Student Nationality',
        'Does the child participate in work?',
        'Does the child have any disability or special need?',
        'fc type',
        'facilitator name',
        'subject taught',
        'date of monitoring',
        'targeted competencies',
        'activities reported',
        'activities reported other',
        'share expectations',
        'share expectations no reason',
        'share expectations other reason',
        'materials needed available',
        'attend lesson',
        'child interact teacher',
        'child interact friends',
        'child clear responses',
        'child ask questions',
        'child acquire competency',
        'child show improvement',
        'child expected work independently',
        'work independently evaluation',
        'complete printed package',
        'sessions participated',
        'not participating reason',
        'E recharge card provided',
        'action to taken',
        'action to taken specify',
        'child needs pss',
        'child cant access resources',
        'homework after lesson',
        'parents supporting student',
        'completed tasks',
        'meet objectives',
        'meet objectives verified',
        'objectives verified specify',
        'additional notes'
    ]

    for col_num in range(len(column_header_fc)):
        wsFC.write(row_num_fc, col_num, column_header_fc[col_num], font_style)
    font_style = xlwt.XFStyle()

    rows_fc = queryset_fc.values_list(

        'enrollment_id',
        'enrollment__partner__name',
        'enrollment__round__name',
        'enrollment__new_registry',
        'enrollment__registration_level',
        'enrollment__student__first_name',
        'enrollment__student__father_name',
        'enrollment__student__last_name',
        'enrollment__student__number',
        'enrollment__student__sex',
        'enrollment__student__nationality__name',
        'enrollment__disability__name_en',
        'enrollment__have_labour_single_selection',
        'fc_type',
        'facilitator_name',
        'subject_taught',
        'date_of_monitoring',
        'targeted_competencies',
        'activities_reported',
        'activities_reported_other',
        'share_expectations',
        'share_expectations_no_reason',
        'share_expectations_other_reason',
        'materials_needed_available',
        'attend_lesson',
        'child_interact_teacher',
        'child_interact_friends',
        'child_clear_responses',
        'child_ask_questions',
        'child_acquire_competency',
        'child_show_improvement',
        'child_expected_work_independently',
        'work_independently_evaluation',
        'complete_printed_package',
        'sessions_participated',
        'not_participating_reason',
        'e_recharge_card_provided',
        'action_to_taken',
        'action_to_taken_specify',
        'child_needs_pss',
        'child_cant_access_resources',
        'homework_after_lesson',
        'parents_supporting_student',
        'completed_tasks',
        'meet_objectives',
        'meet_objectives_verified',
        'objectives_verified_specify',
        'additional_notes'

    )

    for row in rows_fc:
        row_num_fc += 1
        for col_num in range(len(row)):
            wsFC.write(row_num_fc, col_num, row[col_num], font_style)

    wb_student.save(buffer)

    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    response = FileResponse(buffer, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="BLN.xls"'

    return response


def abln_build_xls_extraction(queryset_students, queryset_fc):
    buffer = io.BytesIO()

    # Personnel
    wb_student = xlwt.Workbook(encoding='utf-8', style_compression=2)

    ws = wb_student.add_sheet('Student')

    # Sheet header, first row
    row_num_student = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = [
        'enrollment_id',
        'First time registered?',
        'Partner',
        'CLM Round',
        'Governorate',
        'District',
        'Cadaster',
        'Location',
        'Center',
        'The language supported in the program',
        'Student Address',
        'Registration level',
        'first attendance date',
        'ID number',
        'unique number',
        'First name',
        'Father name',
        'Last name',
        'Mother fullname',
        'Gender',
        'Student Nationality',
        'Student Nationality Specify',
        'Birthday - day',
        'Birthday - month',
        'Birthday - year',
        'P-Code If a child lives in a tent / Brax in a random camp',
        'Does the child have any disability or special need?',
        'Education status',
        'Miss school date',
        'Internal number',
        'RIMS  case number',
        'Source of Identification',
        'Source of Identification Specify',
        'Source of Transportation',
        'What is the educational level of the mother?',
        'What is the educational level of the father?',
        'Phone number',
        'Phone number confirm',
        'phone owner',
        'Second Phone number',
        'Second Phone number confirm',
        'Second phone owner',
        'Main Caregiver',
        'main caregiver nationality',
        'other caregiver relationship',
        'main caregiver nationality Other ',
        'Caretaker first name',
        'Caretaker middle name',
        'Caretaker last name',
        'Caretaker mother name',
        'caretaker birthday year',
        'caretaker birthday month',
        'caretaker birthday day',
        'ID Type',
        'UNHCR case number',
        'UNHCR case number confirm',
        'Parent individual ID',
        'Parent individual ID confirm',
        'Child individual ID',
        'Child individual ID confirm',
        'UNHCR recorded barcode',
        'UNHCR recorded barcode confirm',
        'Parent Lebanese ID number',
        'Parent Lebanese ID number confirm',
        'Child Lebanese ID number',
        'Child Lebanese ID number confirm',
        'Parent Syrian ID number',
        'Parent Syrian ID number confirm',
        'Child Syrian ID number',
        'Child Syrian ID number confirm',
        'Parent Palestinian ID number',
        'Parent Palestinian ID number confirm',
        'Child Palestinian ID number',
        'Child Palestinian ID number confirm',
        'ID number of the Caretaker',
        'ID number of the Caretaker confirm',
        'ID number of the child',
        'ID number of the child confirm',
        'What is the family status of the child?',
        'Does the child have children?',
        'Child number of children',
        'Does the child participate in work?',
        'What is the type of work?',
        'Please specify',
        'How many hours does this child work in a day?',
        'Child weekly income',
        'Level of participation / Absence',
        'The main barriers affecting the daily attendance and performance of the child or drop out of',
        'Please specify',
        'test_done',
        'round_complete',
        'Did the child receive basic stationery?',
        'Did the child benefit from the PSS kit?',
        'Based on the overall score, what is the recommended learning path?',
        'Please specify',
        'Did the child use Akelius program',
        'cp_referral',
        'referal_wash',
        'referal_health',
        'referal_other',
        'referal_other_specify',
        'child_received_books',
        'child_received_printout',
        'child_received_internet',
        'Please enter the number phone calls',
        'Phone call Result of follow up',
        'Please enter the number of house visits',
        'House VisitvResult of follow up',
        'Please enter the number of family visit',
        'Familiy Visit Result of follow up',
        'Parent attended visits',
        'pss session attended',
        'pss session number',
        'pss session modality',
        'pss parent attended',
        'pss parent attended other',
        'Attended covid Session?',
        'Please enter the number of sessions',
        'Please indicate modality',
        'Parent who attended the parents meeting',
        'Please specify',
        'Attended followup Session',
        'Please enter the number of sessions',
        'Please indicate modality',
        'Parent who attended the parents meeting',
        'Please specify',
        'pre test attended arabic',
        'pre test modality arabic',
        'pre test arabic',
        'pre test attended artistic',
        'pre test modality artistic',
        'pre test artistic',
        'pre test attended math',
        'pre test modality math',
        'pre test math',
        'pre test attended social',
        'pre test modality social',
        'pre test social emotional',
        'post test attended arabic',
        'post test modality arabic',
        'post test arabic',
        'post test attended artistic',
        'post test modality artistic',
        'post test artistic',
        'post test attended math',
        'post test modality math',
        'post test math',
        'post test attended social',
        'post test modality social',
        'post test social emotional',
        'owner',
        'modified_by',
    ]

    for col_num in range(len(columns)):
        ws.write(row_num_student, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    qs = queryset_students.extra(select={

        'pre_test_attended_arabic': "pre_test->>'ABLN_ASSESSMENT/attended_arabic'",
        'pre_test_modality_arabic': "pre_test->>'ABLN_ASSESSMENT/modality_arabic'",
        'pre_test_arabic': "pre_test->>'ABLN_ASSESSMENT/arabic'",

        'pre_test_attended_artistic': "pre_test->>'ABLN_ASSESSMENT/attended_artistic'",
        'pre_test_modality_artistic': "pre_test->>'ABLN_ASSESSMENT/modality_artistic'",
        'pre_test_artistic': "pre_test->>'ABLN_ASSESSMENT/artistic'",

        'pre_test_attended_math': "pre_test->>'ABLN_ASSESSMENT/attended_math'",
        'pre_test_modality_math': "pre_test->>'ABLN_ASSESSMENT/modality_math'",
        'pre_test_math': "pre_test->>'ABLN_ASSESSMENT/math'",

        'pre_test_attended_social': "pre_test->>'ABLN_ASSESSMENT/attended_social'",
        'pre_test_modality_social': "pre_test->>'ABLN_ASSESSMENT/modality_social'",
        'pre_test_social_emotional': "pre_test->>'ABLN_ASSESSMENT/social_emotional'",

        'post_test_attended_arabic': "post_test->>'ABLN_ASSESSMENT/attended_arabic'",
        'post_test_modality_arabic': "post_test->>'ABLN_ASSESSMENT/modality_arabic'",
        'post_test_arabic': "post_test->>'ABLN_ASSESSMENT/arabic'",

        'post_test_attended_artistic': "post_test->>'ABLN_ASSESSMENT/attended_artistic'",
        'post_test_modality_artistic': "post_test->>'ABLN_ASSESSMENT/modality_artistic'",
        'post_test_artistic': "post_test->>'ABLN_ASSESSMENT/artistic'",

        'post_test_attended_math': "post_test->>'ABLN_ASSESSMENT/attended_math'",
        'post_test_modality_math': "post_test->>'ABLN_ASSESSMENT/modality_math'",
        'post_test_math': "post_test->>'ABLN_ASSESSMENT/math'",

        'post_test_attended_social': "post_test->>'ABLN_ASSESSMENT/attended_social'",
        'post_test_modality_social': "post_test->>'ABLN_ASSESSMENT/modality_social'",
        'post_test_social_emotional': "post_test->>'ABLN_ASSESSMENT/social_emotional'",
    })

    rows = qs.order_by('id').values_list(
        'id',
        'new_registry',
        'partner__name',
        'round__name',
        'governorate__name_en',
        'district__name_en',
        'cadaster__name_en',
        'location',
        'center__name',
        'language',
        'student__address',
        'registration_level',
        'first_attendance_date',
        'student__id_number',
        'student__number',
        'student__first_name',
        'student__father_name',
        'student__last_name',
        'student__mother_fullname',
        'student__sex',
        'student__nationality__name',
        'other_nationality',
        'student__birthday_day',
        'student__birthday_month',
        'student__birthday_year',
        'student__p_code',
        'disability__name_en',
        'education_status',
        'miss_school_date',
        'internal_number',
        'rims_case_number',
        'source_of_identification',
        'source_of_identification_specify',
        'source_of_transportation',
        'hh_educational_level__name',
        'father_educational_level__name',
        'phone_number',
        'phone_number_confirm',
        'phone_owner',
        'second_phone_number',
        'second_phone_number_confirm',
        'second_phone_owner',
        'main_caregiver',
        'main_caregiver_nationality__name',
        'other_caregiver_relationship',
        'main_caregiver_nationality_other',
        'caretaker_first_name',
        'caretaker_middle_name',
        'caretaker_last_name',
        'caretaker_mother_name',
        'caretaker_birthday_year',
        'caretaker_birthday_month',
        'caretaker_birthday_day',
        'id_type',
        'case_number',
        'case_number_confirm',
        'parent_individual_case_number',
        'parent_individual_case_number_confirm',
        'individual_case_number',
        'individual_case_number_confirm',
        'recorded_number',
        'recorded_number_confirm',
        'parent_national_number',
        'parent_national_number_confirm',
        'national_number',
        'national_number_confirm',
        'parent_syrian_national_number',
        'parent_syrian_national_number_confirm',
        'syrian_national_number',
        'syrian_national_number_confirm',
        'parent_sop_national_number',
        'parent_sop_national_number_confirm',
        'sop_national_number',
        'sop_national_number_confirm',
        'parent_other_number',
        'parent_other_number_confirm',
        'other_number',
        'other_number_confirm',
        'student__family_status',
        'student__have_children',
        'student_number_children',
        'have_labour_single_selection',
        'labours_single_selection',
        'labours_other_specify',
        'labour_hours',
        'labour_weekly_income',
        'participation',
        'barriers_single',
        'barriers_other',
        'test_done',
        'round_complete',
        'basic_stationery',
        'pss_kit',
        'learning_result',
        'learning_result_other',
        'akelius_program',
        'cp_referral',
        'referal_wash',
        'referal_health',
        'referal_other',
        'referal_other_specify',
        'child_received_books',
        'child_received_printout',
        'child_received_internet',
        'phone_call_number',
        'phone_call_follow_up_result',
        'house_visit_number',
        'house_visit_follow_up_result',
        'family_visit_number',
        'family_visit_follow_up_result',
        'parent_attended_visits',
        'pss_session_attended',
        'pss_session_number',
        'pss_session_modality',
        'pss_parent_attended',
        'pss_parent_attended_other',
        'covid_session_attended',
        'covid_session_number',
        'covid_session_modality',
        'covid_parent_attended',
        'covid_parent_attended_other',
        'followup_session_attended',
        'followup_session_number',
        'followup_session_modality',
        'followup_parent_attended',
        'followup_parent_attended_other',
        'pre_test_attended_arabic',
        'pre_test_modality_arabic',
        'pre_test_arabic',
        'pre_test_attended_artistic',
        'pre_test_modality_artistic',
        'pre_test_artistic',
        'pre_test_attended_math',
        'pre_test_modality_math',
        'pre_test_math',
        'pre_test_attended_social',
        'pre_test_modality_social',
        'pre_test_social_emotional',
        'post_test_attended_arabic',
        'post_test_modality_arabic',
        'post_test_arabic',
        'post_test_attended_artistic',
        'post_test_modality_artistic',
        'post_test_artistic',
        'post_test_attended_math',
        'post_test_modality_math',
        'post_test_math',
        'post_test_attended_social',
        'post_test_modality_social',
        'post_test_social_emotional',
        'owner__username',
        'modified_by__username',
    )
    for row in rows:
        row_num_student += 1
        for col_num in range(len(row)):
            ws.write(row_num_student, col_num, row[col_num], font_style)
    # FC
    wsFC = wb_student.add_sheet('FC')
    row_num_fc = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    column_header_fc = [
        'enrollment id',
        'Partner',
        'CLM Round',
        'First time registered?',
        'Registration level',
        'Student first name',
        'Student father name',
        'Student last name',
        'unique number',
        'Gender',
        'Student Nationality',
        'Does the child participate in work?',
        'Does the child have any disability or special need?',
        'fc type',
        'facilitator name',
        'subject taught',
        'date of monitoring',
        'targeted competencies',
        'activities reported',
        'activities reported other',
        'share expectations',
        'share expectations no reason',
        'share expectations other reason',
        'materials needed available',
        'attend lesson',
        'child interact teacher',
        'child interact friends',
        'child clear responses',
        'child ask questions',
        'child acquire competency',
        'child show improvement',
        'child expected work independently',
        'work independently evaluation',
        'complete printed package',
        'sessions participated',
        'not participating reason',
        'E recharge card provided',
        'action to taken',
        'action to taken specify',
        'child needs pss',
        'child cant access resources',
        'homework after lesson',
        'parents supporting student',
        'completed tasks',
        'meet objectives',
        'meet objectives verified',
        'objectives verified specify',
        'additional notes'
    ]

    for col_num in range(len(column_header_fc)):
        wsFC.write(row_num_fc, col_num, column_header_fc[col_num], font_style)
    font_style = xlwt.XFStyle()

    rows_fc = queryset_fc.values_list(

        'enrollment_id',
        'enrollment__partner__name',
        'enrollment__round__name',
        'enrollment__new_registry',
        'enrollment__registration_level',
        'enrollment__student__first_name',
        'enrollment__student__father_name',
        'enrollment__student__last_name',
        'enrollment__student__number',
        'enrollment__student__sex',
        'enrollment__student__nationality__name',
        'enrollment__disability__name_en',
        'enrollment__have_labour_single_selection',
        'fc_type',
        'facilitator_name',
        'subject_taught',
        'date_of_monitoring',
        'targeted_competencies',
        'activities_reported',
        'activities_reported_other',
        'share_expectations',
        'share_expectations_no_reason',
        'share_expectations_other_reason',
        'materials_needed_available',
        'attend_lesson',
        'child_interact_teacher',
        'child_interact_friends',
        'child_clear_responses',
        'child_ask_questions',
        'child_acquire_competency',
        'child_show_improvement',
        'child_expected_work_independently',
        'work_independently_evaluation',
        'complete_printed_package',
        'sessions_participated',
        'not_participating_reason',
        'e_recharge_card_provided',
        'action_to_taken',
        'action_to_taken_specify',
        'child_needs_pss',
        'child_cant_access_resources',
        'homework_after_lesson',
        'parents_supporting_student',
        'completed_tasks',
        'meet_objectives',
        'meet_objectives_verified',
        'objectives_verified_specify',
        'additional_notes'

    )

    for row in rows_fc:
        row_num_fc += 1
        for col_num in range(len(row)):
            wsFC.write(row_num_fc, col_num, row[col_num], font_style)

    wb_student.save(buffer)

    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    response = FileResponse(buffer, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="ABLN.xls"'

    return response


def cbece_build_xls_extraction(queryset_students, queryset_fc):
    buffer = io.BytesIO()

    # Personnel
    wb_student = xlwt.Workbook(encoding='utf-8', style_compression=2)

    ws = wb_student.add_sheet('Student')

    # Sheet header, first row
    row_num_student = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = [
        'enrollment_id',
        'First time registered?',
        'Partner',
        'CLM Round',
        'Governorate',
        'District',
        'Cadaster',
        'Location',
        'Center',
        'The language supported in the program',
        'Student Address',
        'Registration level',
        'first attendance date',
        'ID number',
        'unique number',
        'First name',
        'Father name',
        'Last name',
        'Mother fullname',
        'Gender',
        'Student Nationality',
        'Student Nationality Specify',
        'Birthday - day',
        'Birthday - month',
        'Birthday - year',
        'P-Code If a child lives in a tent / Brax in a random camp',
        'Does the child have any disability or special need?',
        'Education status',
        'Miss school date',
        'Internal number',
        'RIMS  case number',
        'Source of Identification',
        'Source of Identification Specify',
        'Source of Transportation',
        'What is the educational level of the mother?',
        'What is the educational level of the father?',
        'Phone number',
        'Phone number confirm',
        'phone owner',
        'Second Phone number',
        'Second Phone number confirm',
        'Second phone owner',
        'Main Caregiver',
        'main caregiver nationality',
        'other caregiver relationship',
        'Caretaker first name',
        'Caretaker middle name',
        'Caretaker last name',
        'Caretaker mother name',
        'caretaker birthday year',
        'caretaker birthday month',
        'caretaker birthday day',
        'ID Type',
        'UNHCR case number',
        'UNHCR case number confirm',
        'Parent individual ID',
        'Parent individual ID confirm',
        'Child individual ID',
        'Child individual ID confirm',
        'UNHCR recorded barcode',
        'UNHCR recorded barcode confirm',
        'Parent Lebanese ID number',
        'Parent Lebanese ID number confirm',
        'Child Lebanese ID number',
        'Child Lebanese ID number confirm',
        'Parent Syrian ID number',
        'Parent Syrian ID number confirm',
        'Child Syrian ID number',
        'Child Syrian ID number confirm',
        'Parent Palestinian ID number',
        'Parent Palestinian ID number confirm',
        'Child Palestinian ID number',
        'Child Palestinian ID number confirm',
        'ID number of the Caretaker',
        'ID number of the Caretaker confirm',
        'ID number of the child',
        'ID number of the child confirm',
        'What is the family status of the child?',
        'Does the child have children?',
        'Child number of children',
        'Does the child participate in work?',
        'What is the type of work?',
        'Please specify',
        'How many hours does this child work in a day?',
        'Child weekly income',
        'Level of participation / Absence',
        'The main barriers affecting the daily attendance and performance of the child or drop out of',
        'Please specify',
        'test_done',
        'round_complete',
        'Did the child receive basic stationery?',
        'Did the child benefit from the PSS kit?',
        'Based on the overall score, what is the recommended learning path?',
        'Please specify',
        'Did the child use Akelius program',
        'cp_referral',
        'referal_wash',
        'referal_health',
        'referal_other',
        'referal_other_specify',
        'child_received_books',
        'child_received_printout',
        'child_received_internet',
        'Please enter the number phone calls',
        'Phone call Result of follow up',
        'Please enter the number of house visits',
        'House VisitvResult of follow up',
        'Please enter the number of family visit',
        'Familiy Visit Result of follow up',
        'Parent attended visits',
        'pss session attended',
        'pss session number',
        'pss session modality',
        'pss parent attended',
        'pss parent attended other',
        'Attended covid Session?',
        'Please enter the number of sessions',
        'Please indicate modality',
        'Parent who attended the parents meeting',
        'Please specify',
        'Attended followup Session',
        'Please enter the number of sessions',
        'Please indicate modality',
        'Parent who attended the parents meeting',
        'Please specify',
        'pre test attended arabic',
        'pre test modality arabic',
        'pre test arabic',
        'pre test attended english',
        'pre test modality english',
        'pre test english',
        'pre test attended psychomotor',
        'pre test modality psychomotor',
        'pre test psychomotor',
        'pre test attended artistic',
        'pre test modality artistic',
        'pre test artistic',
        'pre test attended math',
        'pre test modality math',
        'pre test math',
        'pre test attended social',
        'pre test modality social',
        'pre test social emotional',
        'pre test attended science',
        'pre test modality science',
        'pre test science',
        'post test attended arabic',
        'post test modality arabic',
        'post test arabic',
        'post test attended english',
        'post test modality english',
        'post test english',
        'post test attended psychomotor',
        'post test modality psychomotor',
        'post test psychomotor',
        'post test attended artistic',
        'post test modality artistic',
        'post test artistic',
        'post test attended math',
        'post test modality math',
        'post test math',
        'post test attended social',
        'post test modality social',
        'post test social emotional',
        'post test attended science',
        'post test modality science',
        'post test science',
        'owner',
        'modified_by',
    ]
    for col_num in range(len(columns)):
        ws.write(row_num_student, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    qs = queryset_students.extra(select={

        'pre_test_attended_arabic': "pre_test->>'CBECE_ASSESSMENT/attended_arabic'",
        'pre_test_modality_arabic': "pre_test->>'CBECE_ASSESSMENT/modality_arabic'",
        'pre_test_arabic': "pre_test->>'CBECE_ASSESSMENT/arabic'",

        'pre_test_attended_english': "pre_test->>'CBECE_ASSESSMENT/attended_english'",
        'pre_test_modality_english': "pre_test->>'CBECE_ASSESSMENT/modality_english'",
        'pre_test_english': "pre_test->>'CBECE_ASSESSMENT/english'",

        'pre_test_attended_psychomotor': "pre_test->>'CBECE_ASSESSMENT/attended_psychomotor'",
        'pre_test_modality_psychomotor': "pre_test->>'CBECE_ASSESSMENT/modality_psychomotor'",
        'pre_test_psychomotor': "pre_test->>'CBECE_ASSESSMENT/psychomotor'",

        'pre_test_attended_math': "pre_test->>'CBECE_ASSESSMENT/attended_math'",
        'pre_test_modality_math': "pre_test->>'CBECE_ASSESSMENT/modality_math'",
        'pre_test_math': "pre_test->>'CBECE_ASSESSMENT/math'",

        'pre_test_attended_social': "pre_test->>'CBECE_ASSESSMENT/attended_social'",
        'pre_test_modality_social': "pre_test->>'CBECE_ASSESSMENT/modality_social'",
        'pre_test_social_emotional': "pre_test->>'CBECE_ASSESSMENT/social_emotional'",

        'pre_test_attended_science': "pre_test->>'CBECE_ASSESSMENT/attended_science'",
        'pre_test_modality_science': "pre_test->>'CBECE_ASSESSMENT/modality_science'",
        'pre_test_science': "pre_test->>'CBECE_ASSESSMENT/science'",

        'pre_test_attended_artistic': "pre_test->>'CBECE_ASSESSMENT/attended_artistic'",
        'pre_test_modality_artistic': "pre_test->>'CBECE_ASSESSMENT/modality_artistic'",
        'pre_test_artistic': "pre_test->>'CBECE_ASSESSMENT/artistic'",

        'post_test_attended_arabic': "post_test->>'CBECE_ASSESSMENT/attended_arabic'",
        'post_test_modality_arabic': "post_test->>'CBECE_ASSESSMENT/modality_arabic'",
        'post_test_arabic': "post_test->>'CBECE_ASSESSMENT/arabic'",

        'post_test_attended_english': "post_test->>'CBECE_ASSESSMENT/attended_english'",
        'post_test_modality_english': "post_test->>'CBECE_ASSESSMENT/modality_english'",
        'post_test_english': "post_test->>'CBECE_ASSESSMENT/english'",

        'post_test_attended_psychomotor': "post_test->>'CBECE_ASSESSMENT/attended_psychomotor'",
        'post_test_modality_psychomotor': "post_test->>'CBECE_ASSESSMENT/modality_psychomotor'",
        'post_test_psychomotor': "post_test->>'CBECE_ASSESSMENT/psychomotor'",

        'post_test_attended_math': "post_test->>'CBECE_ASSESSMENT/attended_math'",
        'post_test_modality_math': "post_test->>'CBECE_ASSESSMENT/modality_math'",
        'post_test_math': "post_test->>'CBECE_ASSESSMENT/math'",

        'post_test_attended_social': "post_test->>'CBECE_ASSESSMENT/attended_social'",
        'post_test_modality_social': "post_test->>'CBECE_ASSESSMENT/modality_social'",
        'post_test_social_emotional': "post_test->>'CBECE_ASSESSMENT/social_emotional'",

        'post_test_attended_science': "post_test->>'CBECE_ASSESSMENT/attended_science'",
        'post_test_modality_science': "post_test->>'CBECE_ASSESSMENT/modality_science'",
        'post_test_science': "post_test->>'CBECE_ASSESSMENT/science'",

        'post_test_attended_artistic': "post_test->>'CBECE_ASSESSMENT/attended_artistic'",
        'post_test_modality_artistic': "post_test->>'CBECE_ASSESSMENT/modality_artistic'",
        'post_test_artistic': "post_test->>'CBECE_ASSESSMENT/artistic'",
    })

    rows = qs.order_by('id').values_list(
        'id',
        'new_registry',
        'partner__name',
        'round__name',
        'governorate__name_en',
        'district__name_en',
        'cadaster__name_en',
        'location',
        'center__name',
        'language',
        'student__address',
        'registration_level',
        'first_attendance_date',
        'student__id_number',
        'student__number',
        'student__first_name',
        'student__father_name',
        'student__last_name',
        'student__mother_fullname',
        'student__sex',
        'student__nationality__name',
        'other_nationality',
        'student__birthday_day',
        'student__birthday_month',
        'student__birthday_year',
        'student__p_code',
        'disability__name_en',
        'education_status',
        'miss_school_date',
        'internal_number',
        'rims_case_number',
        'source_of_identification',
        'source_of_identification_specify',
        'source_of_transportation',
        'hh_educational_level__name',
        'father_educational_level__name',
        'phone_number',
        'phone_number_confirm',
        'phone_owner',
        'second_phone_number',
        'second_phone_number_confirm',
        'second_phone_owner',
        'main_caregiver',
        'main_caregiver_nationality__name',
        'other_caregiver_relationship',
        'caretaker_first_name',
        'caretaker_middle_name',
        'caretaker_last_name',
        'caretaker_mother_name',
        'caretaker_birthday_year',
        'caretaker_birthday_month',
        'caretaker_birthday_day',
        'id_type',
        'case_number',
        'case_number_confirm',
        'parent_individual_case_number',
        'parent_individual_case_number_confirm',
        'individual_case_number',
        'individual_case_number_confirm',
        'recorded_number',
        'recorded_number_confirm',
        'parent_national_number',
        'parent_national_number_confirm',
        'national_number',
        'national_number_confirm',
        'parent_syrian_national_number',
        'parent_syrian_national_number_confirm',
        'syrian_national_number',
        'syrian_national_number_confirm',
        'parent_sop_national_number',
        'parent_sop_national_number_confirm',
        'sop_national_number',
        'sop_national_number_confirm',
        'parent_other_number',
        'parent_other_number_confirm',
        'other_number',
        'other_number_confirm',
        'student__family_status',
        'student__have_children',
        'student_number_children',
        'have_labour_single_selection',
        'labours_single_selection',
        'labours_other_specify',
        'labour_hours',
        'labour_weekly_income',
        'participation',
        'barriers_single',
        'barriers_other',
        'test_done',
        'round_complete',
        'basic_stationery',
        'pss_kit',
        'learning_result',
        'learning_result_other',
        'akelius_program',
        'cp_referral',
        'referal_wash',
        'referal_health',
        'referal_other',
        'referal_other_specify',
        'child_received_books',
        'child_received_printout',
        'child_received_internet',
        'phone_call_number',
        'phone_call_follow_up_result',
        'house_visit_number',
        'house_visit_follow_up_result',
        'family_visit_number',
        'family_visit_follow_up_result',
        'parent_attended_visits',
        'pss_session_attended',
        'pss_session_number',
        'pss_session_modality',
        'pss_parent_attended',
        'pss_parent_attended_other',
        'covid_session_attended',
        'covid_session_number',
        'covid_session_modality',
        'covid_parent_attended',
        'covid_parent_attended_other',
        'followup_session_attended',
        'followup_session_number',
        'followup_session_modality',
        'followup_parent_attended',
        'followup_parent_attended_other',
        'pre_test_attended_arabic',
        'pre_test_modality_arabic',
        'pre_test_arabic',
        'pre_test_attended_english',
        'pre_test_modality_english',
        'pre_test_english',
        'pre_test_attended_psychomotor',
        'pre_test_modality_psychomotor',
        'pre_test_psychomotor',
        'pre_test_attended_artistic',
        'pre_test_modality_artistic',
        'pre_test_artistic',
        'pre_test_attended_math',
        'pre_test_modality_math',
        'pre_test_math',
        'pre_test_attended_social',
        'pre_test_modality_social',
        'pre_test_social_emotional',
        'pre_test_attended_science',
        'pre_test_modality_science',
        'pre_test_science',
        'post_test_attended_arabic',
        'post_test_modality_arabic',
        'post_test_arabic',
        'post_test_attended_english',
        'post_test_modality_english',
        'post_test_english',
        'post_test_attended_psychomotor',
        'post_test_modality_psychomotor',
        'post_test_psychomotor',
        'post_test_attended_artistic',
        'post_test_modality_artistic',
        'post_test_artistic',
        'post_test_attended_math',
        'post_test_modality_math',
        'post_test_math',
        'post_test_attended_social',
        'post_test_modality_social',
        'post_test_social_emotional',
        'post_test_attended_science',
        'post_test_modality_science',
        'post_test_science',
        'owner__username',
        'modified_by__username',
    )
    for row in rows:
        row_num_student += 1
        for col_num in range(len(row)):
            ws.write(row_num_student, col_num, row[col_num], font_style)
    # FC
    wsFC = wb_student.add_sheet('FC')
    row_num_fc = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    column_header_fc = [
        'enrollment id',
        'Partner',
        'CLM Round',
        'First time registered?',
        'Registration level',
        'Student first name',
        'Student father name',
        'Student last name',
        'unique number',
        'Gender',
        'Student Nationality',
        'Does the child participate in work?',
        'Does the child have any disability or special need?',
        'fc type',
        'facilitator name',
        'subject taught',
        'date of monitoring',
        'targeted competencies',
        'activities reported',
        'activities reported other',
        'share expectations',
        'share expectations no reason',
        'share expectations other reason',
        'materials needed available',
        'attend lesson',
        'child interact teacher',
        'child interact friends',
        'child clear responses',
        'child ask questions',
        'child acquire competency',
        'child show improvement',
        'child expected work independently',
        'work independently evaluation',
        'complete printed package',
        'sessions participated',
        'not participating reason',
        'E recharge card provided',
        'action to taken',
        'action to taken specify',
        'child needs pss',
        'child cant access resources',
        'homework after lesson',
        'parents supporting student',
        'completed tasks',
        'meet objectives',
        'meet objectives verified',
        'objectives verified specify',
        'additional notes'
    ]

    for col_num in range(len(column_header_fc)):
        wsFC.write(row_num_fc, col_num, column_header_fc[col_num], font_style)
    font_style = xlwt.XFStyle()

    rows_fc = queryset_fc.values_list(

        'enrollment_id',
        'enrollment__partner__name',
        'enrollment__round__name',
        'enrollment__new_registry',
        'enrollment__registration_level',
        'enrollment__student__first_name',
        'enrollment__student__father_name',
        'enrollment__student__last_name',
        'enrollment__student__number',
        'enrollment__student__sex',
        'enrollment__student__nationality__name',
        'enrollment__disability__name_en',
        'enrollment__have_labour_single_selection',
        'fc_type',
        'facilitator_name',
        'subject_taught',
        'date_of_monitoring',
        'targeted_competencies',
        'activities_reported',
        'activities_reported_other',
        'share_expectations',
        'share_expectations_no_reason',
        'share_expectations_other_reason',
        'materials_needed_available',
        'attend_lesson',
        'child_interact_teacher',
        'child_interact_friends',
        'child_clear_responses',
        'child_ask_questions',
        'child_acquire_competency',
        'child_show_improvement',
        'child_expected_work_independently',
        'work_independently_evaluation',
        'complete_printed_package',
        'sessions_participated',
        'not_participating_reason',
        'e_recharge_card_provided',
        'action_to_taken',
        'action_to_taken_specify',
        'child_needs_pss',
        'child_cant_access_resources',
        'homework_after_lesson',
        'parents_supporting_student',
        'completed_tasks',
        'meet_objectives',
        'meet_objectives_verified',
        'objectives_verified_specify',
        'additional_notes'

    )

    for row in rows_fc:
        row_num_fc += 1
        for col_num in range(len(row)):
            wsFC.write(row_num_fc, col_num, row[col_num], font_style)

    wb_student.save(buffer)

    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    response = FileResponse(buffer, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="CBECE.xls"'

    return response


def rs_build_xls_extraction(queryset_students, queryset_fc):
    buffer = io.BytesIO()

    wb = Workbook()

    ws_student = wb['Sheet']
    ws_student.title = 'Student'

    # Sheet header, first row
    row_num_student = 0

    font_style = Font(bold=True)

    columns = [
        'enrollment_id',
        'First time registered?',
        'Partner',
        'CLM Round',
        'Governorate',
        'District',
        'Cadaster',
        'Location',
        'Center',
        'The language supported in the program',
        'Grade of registration',
        'Student Address',
        'Registration level',
        'first attendance date',
        'School of Enrollment',
        'shift',
        'ID number',
        'unique number',
        'First name',
        'Father name',
        'Last name',
        'Mother fullname',
        'Gender',
        'Student Nationality',
        'Student Nationality Specify',
        'Birthday - day',
        'Birthday - month',
        'Birthday - year',
        'P-Code If a child lives in a tent / Brax in a random camp',
        'Does the child have any disability or special need?',
        'What was the child education level when first joining formal education in lebanon',
        'From where did the child first come to join  FE',
        'Education status',
        'Miss school?',
        'Miss school date',
        'Internal number',
        'RIMS  case number',
        'Source of Identification',
        'Source of Identification Specify',
        'Source of Transportation',
        'What is the educational level of the mother?',
        'What is the educational level of the father?',
        'Phone number',
        'Phone number confirm',
        'phone owner',
        'Second Phone number',
        'Second Phone number confirm',
        'Second phone owner',
        'Main Caregiver',
        'main caregiver nationality',
        'other caregiver relationship',
        'Caretaker first name',
        'Caretaker middle name',
        'Caretaker last name',
        'Caretaker mother name',
        'caretaker birthday year',
        'caretaker birthday month',
        'caretaker birthday day',
        'ID Type',
        'UNHCR case number',
        'UNHCR case number confirm',
        'Parent individual ID',
        'Parent individual ID confirm',
        'Child individual ID',
        'Child individual ID confirm',
        'UNHCR recorded barcode',
        'UNHCR recorded barcode confirm',
        'Parent Lebanese ID number',
        'Parent Lebanese ID number confirm',
        'Child Lebanese ID number',
        'Child Lebanese ID number confirm',
        'Parent Syrian ID number',
        'Parent Syrian ID number confirm',
        'Child Syrian ID number',
        'Child Syrian ID number confirm',
        'Parent Palestinian ID number',
        'Parent Palestinian ID number confirm',
        'Child Palestinian ID number',
        'Child Palestinian ID number confirm',
        'ID number of the Caretaker',
        'ID number of the Caretaker confirm',
        'ID number of the child',
        'ID number of the child confirm',
        'What is the family status of the child?',
        'Does the child have children?',
        'Child number of children',
        'Does the child participate in work?',
        'What is the type of work?',
        'Please specify',
        'How many hours does this child work in a day?',
        'Child weekly income',
        'Level of participation / Absence',
        'The main barriers affecting the daily attendance and performance of the child or drop out of',
        'Please specify',
        'test_done',
        'round_complete',
        'Did the child receive basic stationery?',
        'Did the child benefit from the PSS kit?',
        'Based on the overall score, what is the recommended learning path?',
        'Please specify',
        'Did the child use Akelius program',
        'cp_referral',
        'referal_wash',
        'referal_health',
        'referal_other',
        'referal_other_specify',
        'child_received_books',
        'child_received_printout',
        'child_received_internet',
        'Please enter the number phone calls',
        'Phone call Result of follow up',
        'Please enter the number of house visits',
        'House VisitvResult of follow up',
        'Please enter the number of family visit',
        'Familiy Visit Result of follow up',
        'Parent attended visits',
        'pss session attended',
        'pss session number',
        'pss session modality',
        'pss parent attended',
        'pss parent attended other',
        'Attended covid Session?',
        'Please enter the number of sessions',
        'Please indicate modality',
        'Parent who attended the parents meeting',
        'Please specify',
        'Attended followup Session',
        'Please enter the number of sessions',
        'Please indicate modality',
        'Parent who attended the parents meeting',
        'Please specify',
        'pre test attended arabic',
        'pre test modality arabic',
        'pre test arabic',
        'pre test attended english',
        'pre test modality english',
        'pre test english',
        'pre test attended math',
        'pre test modality math',
        'pre test math',
        'pre test attended science',
        'pre test modality science',
        'pre test science',
        'pre test attended biology',
        'pre test modality biology',
        'pre test biology',
        'pre test attended physics',
        'pre test modality physics',
        'pre test physics',
        'pre test attended chemistry',
        'pre test modality chemistry',
        'pre test chemistry',
        'post test attended arabic',
        'post test modality arabic',
        'post test arabic',
        'post test attended english',
        'post test modality english',
        'post test english',
        'post test attended math',
        'post test modality math',
        'post test math',
        'post test attended science',
        'post test modality science',
        'post test science',
        'post test attended biology',
        'post test modality biology',
        'post test biology',
        'post test attended physics',
        'post test modality physics',
        'post test physics',
        'post test attended chemistry',
        'post test modality chemistry',
        'post test chemistry',
        'owner',
        'modified_by',
    ]

    for col_num in range(len(columns)):
        currentCell = ws_student.cell(row=row_num_student + 1, column=col_num + 1, value=columns[col_num])
        currentCell.font = font_style

    qs = queryset_students.extra(select={

        'pre_test_attended_arabic': "pre_test->>'RS_ASSESSMENT/attended_arabic'",
        'pre_test_modality_arabic': "pre_test->>'RS_ASSESSMENT/modality_arabic'",
        'pre_test_arabic': "pre_test->>'RS_ASSESSMENT/arabic'",

        'pre_test_attended_english': "pre_test->>'RS_ASSESSMENT/attended_english'",
        'pre_test_modality_english': "pre_test->>'RS_ASSESSMENT/modality_english'",
        'pre_test_english': "pre_test->>'RS_ASSESSMENT/english'",

        'pre_test_attended_math': "pre_test->>'RS_ASSESSMENT/attended_math'",
        'pre_test_modality_math': "pre_test->>'RS_ASSESSMENT/modality_math'",
        'pre_test_math': "pre_test->>'RS_ASSESSMENT/math'",

        'pre_test_attended_science': "pre_test->>'RS_ASSESSMENT/attended_science'",
        'pre_test_modality_science': "pre_test->>'RS_ASSESSMENT/modality_science'",
        'pre_test_science': "pre_test->>'RS_ASSESSMENT/science'",

        'pre_test_attended_biology': "pre_test->>'RS_ASSESSMENT/attended_biology'",
        'pre_test_modality_biology': "pre_test->>'RS_ASSESSMENT/modality_biology'",
        'pre_test_biology': "pre_test->>'RS_ASSESSMENT/biology'",

        'pre_test_attended_physics': "pre_test->>'RS_ASSESSMENT/attended_physics'",
        'pre_test_modality_physics': "pre_test->>'RS_ASSESSMENT/modality_physics'",
        'pre_test_physics': "pre_test->>'RS_ASSESSMENT/physics'",

        'pre_test_attended_chemistry': "pre_test->>'RS_ASSESSMENT/attended_chemistry'",
        'pre_test_modality_chemistry': "pre_test->>'RS_ASSESSMENT/modality_chemistry'",
        'pre_test_chemistry': "pre_test->>'RS_ASSESSMENT/chemistry'",

        'post_test_attended_arabic': "post_test->>'RS_ASSESSMENT/attended_arabic'",
        'post_test_modality_arabic': "post_test->>'RS_ASSESSMENT/modality_arabic'",
        'post_test_arabic': "post_test->>'RS_ASSESSMENT/arabic'",

        'post_test_attended_english': "post_test->>'RS_ASSESSMENT/attended_english'",
        'post_test_modality_english': "post_test->>'RS_ASSESSMENT/modality_english'",
        'post_test_english': "post_test->>'RS_ASSESSMENT/english'",

        'post_test_attended_math': "post_test->>'RS_ASSESSMENT/attended_math'",
        'post_test_modality_math': "post_test->>'RS_ASSESSMENT/modality_math'",
        'post_test_math': "post_test->>'RS_ASSESSMENT/math'",

        'post_test_attended_science': "post_test->>'RS_ASSESSMENT/attended_science'",
        'post_test_modality_science': "post_test->>'RS_ASSESSMENT/modality_science'",
        'post_test_science': "post_test->>'RS_ASSESSMENT/science'",

        'post_test_attended_biology': "post_test->>'RS_ASSESSMENT/attended_biology'",
        'post_test_modality_biology': "post_test->>'RS_ASSESSMENT/modality_biology'",
        'post_test_biology': "post_test->>'RS_ASSESSMENT/biology'",

        'post_test_attended_physics': "post_test->>'RS_ASSESSMENT/attended_physics'",
        'post_test_modality_physics': "post_test->>'RS_ASSESSMENT/modality_physics'",
        'post_test_physics': "post_test->>'RS_ASSESSMENT/physics'",

        'post_test_attended_chemistry': "post_test->>'RS_ASSESSMENT/attended_chemistry'",
        'post_test_modality_chemistry': "post_test->>'RS_ASSESSMENT/modality_chemistry'",
        'post_test_chemistry': "post_test->>'RS_ASSESSMENT/chemistry'",

    })

    rows = qs.order_by('id').values_list(
        'id',
        'new_registry',
        'partner__name',
        'round__name',
        'governorate__name_en',
        'district__name_en',
        'cadaster__name_en',
        'location',
        'center__name',
        'language',
        'grade_registration',
        'student__address',
        'registration_level',
        'first_attendance_date',
        'registered_in_school',
        'shift',
        'student__id_number',
        'student__number',
        'student__first_name',
        'student__father_name',
        'student__last_name',
        'student__mother_fullname',
        'student__sex',
        'student__nationality__name',
        'other_nationality',
        'student__birthday_day',
        'student__birthday_month',
        'student__birthday_year',
        'student__p_code',
        'disability__name_en',
        'grade_level',
        'source_join_fe',
        'education_status',
        'miss_school',
        'miss_school_date',
        'internal_number',
        'rims_case_number',
        'source_of_identification',
        'source_of_identification_specify',
        'source_of_transportation',
        'hh_educational_level__name',
        'father_educational_level__name',
        'phone_number',
        'phone_number_confirm',
        'phone_owner',
        'second_phone_number',
        'second_phone_number_confirm',
        'second_phone_owner',
        'main_caregiver',
        'main_caregiver_nationality__name',
        'other_caregiver_relationship',
        'caretaker_first_name',
        'caretaker_middle_name',
        'caretaker_last_name',
        'caretaker_mother_name',
        'caretaker_birthday_year',
        'caretaker_birthday_month',
        'caretaker_birthday_day',
        'id_type',
        'case_number',
        'case_number_confirm',
        'parent_individual_case_number',
        'parent_individual_case_number_confirm',
        'individual_case_number',
        'individual_case_number_confirm',
        'recorded_number',
        'recorded_number_confirm',
        'parent_national_number',
        'parent_national_number_confirm',
        'national_number',
        'national_number_confirm',
        'parent_syrian_national_number',
        'parent_syrian_national_number_confirm',
        'syrian_national_number',
        'syrian_national_number_confirm',
        'parent_sop_national_number',
        'parent_sop_national_number_confirm',
        'sop_national_number',
        'sop_national_number_confirm',
        'parent_other_number',
        'parent_other_number_confirm',
        'other_number',
        'other_number_confirm',
        'student__family_status',
        'student__have_children',
        'student_number_children',
        'have_labour_single_selection',
        'labours_single_selection',
        'labours_other_specify',
        'labour_hours',
        'labour_weekly_income',
        'participation',
        'barriers_single',
        'barriers_other',
        'test_done',
        'round_complete',
        'basic_stationery',
        'pss_kit',
        'learning_result',
        'learning_result_other',
        'akelius_program',
        'cp_referral',
        'referal_wash',
        'referal_health',
        'referal_other',
        'referal_other_specify',
        'child_received_books',
        'child_received_printout',
        'child_received_internet',
        'phone_call_number',
        'phone_call_follow_up_result',
        'house_visit_number',
        'house_visit_follow_up_result',
        'family_visit_number',
        'family_visit_follow_up_result',
        'parent_attended_visits',
        'pss_session_attended',
        'pss_session_number',
        'pss_session_modality',
        'pss_parent_attended',
        'pss_parent_attended_other',
        'covid_session_attended',
        'covid_session_number',
        'covid_session_modality',
        'covid_parent_attended',
        'covid_parent_attended_other',
        'followup_session_attended',
        'followup_session_number',
        'followup_session_modality',
        'followup_parent_attended',
        'followup_parent_attended_other',
        'pre_test_attended_arabic',
        'pre_test_modality_arabic',
        'pre_test_arabic',
        'pre_test_attended_english',
        'pre_test_modality_english',
        'pre_test_english',
        'pre_test_attended_math',
        'pre_test_modality_math',
        'pre_test_math',
        'pre_test_attended_science',
        'pre_test_modality_science',
        'pre_test_science',
        'pre_test_attended_biology',
        'pre_test_modality_biology',
        'pre_test_biology',
        'pre_test_attended_physics',
        'pre_test_modality_physics',
        'pre_test_physics',
        'pre_test_attended_chemistry',
        'pre_test_modality_chemistry',
        'pre_test_chemistry',
        'post_test_attended_arabic',
        'post_test_modality_arabic',
        'post_test_arabic',
        'post_test_attended_english',
        'post_test_modality_english',
        'post_test_english',
        'post_test_attended_math',
        'post_test_modality_math',
        'post_test_math',
        'post_test_attended_science',
        'post_test_modality_science',
        'post_test_science',
        'post_test_attended_biology',
        'post_test_modality_biology',
        'post_test_biology',
        'post_test_attended_physics',
        'post_test_modality_physics',
        'post_test_physics',
        'post_test_attended_chemistry',
        'post_test_modality_chemistry',
        'post_test_chemistry',
        'owner__username',
        'modified_by__username',
    )
    # [:5000]

    for row in rows:
        row_num_student += 1
        for col_num in range(len(row)):
            cellValue = row[col_num]
            cellTextValue = cellValue
            if isinstance(cellValue, str):
                cellTextValue = cellValue.encode("utf8")
            if type(cellValue) in [int]:
                cellTextValue = str(cellValue).encode("utf8")
            if type(cellValue) in [unicode, date]:
                cellTextValue = cellValue
            if type(cellValue) in [list]:
                cellTextValue = listToString(cellValue)
            currentCell = ws_student.cell(row=row_num_student + 1, column=col_num + 1, value=cellTextValue)
            currentCell.font = font_style
    #get partners
    rows_partners = qs.distinct('partner_id').order_by('partner_id').values_list(
        'partner__id',
        'partner__name',
    )
    for partner in rows_partners:
        partner_id = partner[0]
        partner_name = str(partner[1]).encode("utf8")

        ws_fc = wb.create_sheet('FC - ' + partner_name)
        # Sheet header, first row
        row_num_fc = 0
        columns_fc = [
                'enrollment id',
                'Partner',
                'CLM Round',
                'First time registered?',
                'Registration level',
                'Student first name',
                'Student father name',
                'Student last name',
                'unique number',
                'Gender',
                'Student Nationality',
                'Does the child participate in work?',
                'Does the child have any disability or special need?',
                'fc type',
                'facilitator name',
                'subject taught',
                'date of monitoring',
                'targeted competencies',
                'activities reported',
                'activities reported other',
                'share expectations',
                'share expectations no reason',
                'share expectations other reason',
                'materials needed available',
                'attend lesson',
                'child interact teacher',
                'child interact friends',
                'child clear responses',
                'child ask questions',
                'child acquire competency',
                'child show improvement',
                'child expected work independently',
                'work independently evaluation',
                'complete printed package',
                'sessions participated',
                'not participating reason',
                'E recharge card provided',
                'action to taken',
                'action to taken specify',
                'child needs pss',
                'child cant access resources',
                'homework after lesson',
                'parents supporting student',
                'completed tasks',
                'meet objectives',
                'meet objectives verified',
                'objectives verified specify',
                'additional notes'
            ]

        for col_num in range(len(columns_fc)):
            currentCell = ws_fc.cell(row=row_num_fc + 1, column=col_num + 1, value=columns_fc[col_num])
            currentCell.font = font_style

        rows_fc = queryset_fc.filter(enrollment__partner=partner_id).values_list(
            'enrollment_id',
            'enrollment__partner__name',
            'enrollment__round__name',
            'enrollment__new_registry',
            'enrollment__registration_level',
            'enrollment__student__first_name',
            'enrollment__student__father_name',
            'enrollment__student__last_name',
            'enrollment__student__number',
            'enrollment__student__sex',
            'enrollment__student__nationality__name',
            'enrollment__disability__name_en',
            'enrollment__have_labour_single_selection',
            'fc_type',
            'facilitator_name',
            'subject_taught',
            'date_of_monitoring',
            'targeted_competencies',
            'activities_reported',
            'activities_reported_other',
            'share_expectations',
            'share_expectations_no_reason',
            'share_expectations_other_reason',
            'materials_needed_available',
            'attend_lesson',
            'child_interact_teacher',
            'child_interact_friends',
            'child_clear_responses',
            'child_ask_questions',
            'child_acquire_competency',
            'child_show_improvement',
            'child_expected_work_independently',
            'work_independently_evaluation',
            'complete_printed_package',
            'sessions_participated',
            'not_participating_reason',
            'e_recharge_card_provided',
            'action_to_taken',
            'action_to_taken_specify',
            'child_needs_pss',
            'child_cant_access_resources',
            'homework_after_lesson',
            'parents_supporting_student',
            'completed_tasks',
            'meet_objectives',
            'meet_objectives_verified',
            'objectives_verified_specify',
            'additional_notes'
        )
        # [:5000]

        for row in rows_fc:
            row_num_fc += 1
            for col_num in range(len(row)):
                cellValue = row[col_num]
                cellTextValue = cellValue
                if isinstance(cellValue, str):
                    cellTextValue = cellValue.encode("utf8")
                if type(cellValue) in [int]:
                    cellTextValue = str(cellValue).encode("utf8")
                if type(cellValue) in [unicode, date]:
                    cellTextValue = cellValue
                if type(cellValue) in [list]:
                    cellTextValue = listToString(cellValue)
                currentCell = ws_fc.cell(row=row_num_fc + 1, column=col_num + 1, value=cellTextValue)
                currentCell.font = font_style


    # Save the file
    wb.save(buffer)

    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    response = FileResponse(buffer, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="RS.xlsx"'

    return response


def outreach_build_xls_extraction(queryset_students):
    buffer = io.BytesIO()

    # Personnel
    wb_student = xlwt.Workbook(encoding='utf-8', style_compression=2)

    ws = wb_student.add_sheet('Student')

    # Sheet header, first row
    row_num_student = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = [
        'enrollment_id',
        'Partner',
        'CLM Round',
        'Governorate',
        'District',
        'Cadaster',
        'Location',
        'Center',
        'The language supported in the program',
        'Student Address',
        'Registration level',
        'First attendance date',
        'Unique number',
        'First name',
        'Father name',
        'Last name',
        'Mother fullname',
        'Gender',
        'Student Nationality',
        'Student Nationality Specify',
        'Birthday - day',
        'Birthday - month',
        'Birthday - year',
        'P-Code If a child lives in a tent / Brax in a random camp',
        'Does the child have any disability or special need?',
        'Education status',
        'Miss school date',
        'Internal number',
        'Source of Identification',
        'Source of Identification Specify',
        'What is the educational level of the mother?',
        'What is the educational level of the father?',
        'Phone number',
        'Phone number confirm',
        'phone owner',
        'Second Phone number',
        'Second Phone number confirm',
        'Second phone owner',
        'Main Caregiver',
        'main caregiver nationality',
        'other caregiver relationship',
        'main caregiver nationality Other ',
        'Caretaker first name',
        'Caretaker middle name',
        'Caretaker last name',
        'Caretaker mother name',
        'caretaker birthday year',
        'caretaker birthday month',
        'caretaker birthday day',
        'ID Type',
        'UNHCR case number',
        'UNHCR case number confirm',
        'Parent individual ID',
        'Parent individual ID confirm',
        'Child individual ID',
        'Child individual ID confirm',
        'UNHCR recorded barcode',
        'UNHCR recorded barcode confirm',
        'Parent Lebanese ID number',
        'Parent Lebanese ID number confirm',
        'Child Lebanese ID number',
        'Child Lebanese ID number confirm',
        'Parent Syrian ID number',
        'Parent Syrian ID number confirm',
        'Child Syrian ID number',
        'Child Syrian ID number confirm',
        'Parent Palestinian ID number',
        'Parent Palestinian ID number confirm',
        'Child Palestinian ID number',
        'Child Palestinian ID number confirm',
        'ID number of the Caretaker',
        'ID number of the Caretaker confirm',
        'ID number of the child',
        'ID number of the child confirm',
        'owner',
        'modified_by',
    ]

    for col_num in range(len(columns)):
        ws.write(row_num_student, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    rows = queryset_students.order_by('id').values_list(
        'id',
        'partner__name',
        'round__name',
        'governorate__name_en',
        'district__name_en',
        'cadaster__name_en',
        'location',
        'center__name',
        'language',
        'student__address',
        'registration_level',
        'first_attendance_date',
        'student__number',
        'student__first_name',
        'student__father_name',
        'student__last_name',
        'student__mother_fullname',
        'student__sex',
        'student__nationality__name',
        'other_nationality',
        'student__birthday_day',
        'student__birthday_month',
        'student__birthday_year',
        'student__p_code',
        'disability__name_en',
        'education_status',
        'miss_school_date',
        'internal_number',
        'source_of_identification',
        'source_of_identification_specify',
        'hh_educational_level__name',
        'father_educational_level__name',
        'phone_number',
        'phone_number_confirm',
        'phone_owner',
        'second_phone_number',
        'second_phone_number_confirm',
        'second_phone_owner',
        'main_caregiver',
        'main_caregiver_nationality__name',
        'other_caregiver_relationship',
        'main_caregiver_nationality_other',
        'caretaker_first_name',
        'caretaker_middle_name',
        'caretaker_last_name',
        'caretaker_mother_name',
        'caretaker_birthday_year',
        'caretaker_birthday_month',
        'caretaker_birthday_day',
        'id_type',
        'case_number',
        'case_number_confirm',
        'parent_individual_case_number',
        'parent_individual_case_number_confirm',
        'individual_case_number',
        'individual_case_number_confirm',
        'recorded_number',
        'recorded_number_confirm',
        'parent_national_number',
        'parent_national_number_confirm',
        'national_number',
        'national_number_confirm',
        'parent_syrian_national_number',
        'parent_syrian_national_number_confirm',
        'syrian_national_number',
        'syrian_national_number_confirm',
        'parent_sop_national_number',
        'parent_sop_national_number_confirm',
        'sop_national_number',
        'sop_national_number_confirm',
        'parent_other_number',
        'parent_other_number_confirm',
        'other_number',
        'other_number_confirm',
        'owner__username',
        'modified_by__username',
    )
    for row in rows:
        row_num_student += 1
        for col_num in range(len(row)):
            ws.write(row_num_student, col_num, row[col_num], font_style)

    wb_student.save(buffer)

    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    response = FileResponse(buffer, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Outreach.xls"'

    return response


def listToString(s):
    # initialize an empty string
    str1 = ""

    # traverse in the string
    for ele in s:
        if str1 == "":
            str1 += ele
        else:
            str1 += "," + ele

        # return string
    return str1


def build_xls_extraction_horizental(queryset_students, queryset_fc):
    buffer = io.BytesIO()

    # Personnel
    wb_student = xlwt.Workbook(encoding='utf-8', style_compression=2)

    ws = wb_student.add_sheet('Student')

    # Sheet header, first row
    row_num_student = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = [
        'enrollment_id',
        'First time registered?',
        'Partner',
        'CLM Round',
        'Governorate',
        'District',
        'Cadaster',
        'Location',
        'Center',
        'The language supported in the program',
        'Student Address',
        'Registration level',
        'first attendance date',
        'ID number',
        'unique number',
        'First name',
        'Father name',
        'Last name',
        'Mother fullname',
        'Sex',
        'Student Nationality',
        'Student Nationality Specify',
        'Birthday - day',
        'Birthday - month',
        'Birthday - year',
        'P-Code If a child lives in a tent / Brax in a random camp',
        'Does the child have any disability or special need?',
        'Education status',
        'Miss school date',
        'Internal number',
        'RIMS  case number',
        'Source of Identification',
        'Source of Identification Specify',
        'Source of Transportation',
        'What is the educational level of the mother?',
        'What is the educational level of the father?',
        'Phone number',
        'Phone number confirm',
        'phone owner',
        'Second Phone number',
        'Second Phone number confirm',
        'Second phone owner',
        'Main Caregiver',
        'main caregiver nationality',
        'other caregiver relationship',
        'main caregiver nationality Other ',
        'Caretaker first name',
        'Caretaker middle name',
        'Caretaker last name',
        'Caretaker mother name',
        'caretaker birthday year',
        'caretaker birthday month',
        'caretaker birthday day',
        'ID Type',
        'UNHCR case number',
        'UNHCR case number confirm',
        'Parent individual ID',
        'Parent individual ID confirm',
        'Child individual ID',
        'Child individual ID confirm',
        'UNHCR recorded barcode',
        'UNHCR recorded barcode confirm',
        'Parent Lebanese ID number',
        'Parent Lebanese ID number confirm',
        'Child Lebanese ID number',
        'Child Lebanese ID number confirm',
        'Parent Syrian ID number',
        'Parent Syrian ID number confirm',
        'Child Syrian ID number',
        'Child Syrian ID number confirm',
        'Parent Palestinian ID number',
        'Parent Palestinian ID number confirm',
        'Child Palestinian ID number',
        'Child Palestinian ID number confirm',
        'ID number of the Caretaker',
        'ID number of the Caretaker confirm',
        'ID number of the child',
        'ID number of the child confirm',
        'What is the family status of the child?',
        'Does the child have children?',
        'Child number of children',
        'Does the child participate in work?',
        'What is the type of work?',
        'Please specify',
        'How many hours does this child work in a day?',
        'Child weekly income',
        'Level of participation / Absence',
        'The main barriers affecting the daily attendance and performance of the child or drop out of',
        'Please specify',
        'test_done',
        'round_complete',
        'Did the child receive basic stationery?',
        'Did the child benefit from the PSS kit?',
        'Based on the overall score, what is the recommended learning path?',
        'Please specify',
        'Did the child use Akelius program',
        'cp_referral',
        'referal_wash',
        'referal_health',
        'referal_other',
        'referal_other_specify',
        'child_received_books',
        'child_received_printout',
        'child_received_internet',
        'Please enter the number phone calls',
        'Phone call Result of follow up',
        'Please enter the number of house visits',
        'House VisitvResult of follow up',
        'Please enter the number of family visit',
        'Familiy Visit Result of follow up',
        'Parent attended visits',
        'pss session attended',
        'pss session number',
        'pss session modality',
        'pss parent attended',
        'pss parent attended other',
        'Attended covid Session?',
        'Please enter the number of sessions',
        'Please indicate modality',
        'Parent who attended the parents meeting',
        'Please specify',
        'Attended followup Session',
        'Please enter the number of sessions',
        'Please indicate modality',
        'Parent who attended the parents meeting',
        'Please specify',
        'pre test attended arabic',
        'pre test modality arabic',
        'pre test arabic',
        'pre test attended english',
        'pre test modality english',
        'pre test english',
        'pre test attended psychomotor',
        'pre test modality psychomotor',
        'pre test psychomotor',
        'pre test attended artistic',
        'pre test modality artistic',
        'pre test artistic',
        'pre test attended math',
        'pre test modality math',
        'pre test math',
        'pre test attended social',
        'pre test modality social',
        'pre test social emotional',
        'post test attended arabic',
        'post test modality arabic',
        'post test arabic',
        'post test attended english',
        'post test modality english',
        'post test english',
        'post test attended psychomotor',
        'post test modality psychomotor',
        'post test psychomotor',
        'post test attended artistic',
        'post test modality artistic',
        'post test artistic',
        'post test attended math',
        'post test modality math',
        'post test math',
        'post test attended social',
        'post test modality social',
        'post test social emotional',
        'Was the child involved in remote learning?',
        'what other reasons for this child not being engaged?',
        'reasons not engaged other',
        'Does the family have reliable internet service in their area during remote learning?',
        'Did both girls and boys in the same family participate in the class and have access to the phone/',
        'Explain',
        'Frequency of Child Engagement in remote learning?',
        'How well did the child meet the learning outcomes?',
        'How do you rate the parents learning support provided to the child through this Remote',
        'Has the child directly been reached with awareness messaging on Covid-19 and prevention measures?',
        'How often?',
        'Has the parents directly been reached with awareness messaging on Covid-19 and prevention',
        'How often?',
        'Was any follow-up done to ensure messages were well received, understood and adopted?',
        'With who child and/or caregiver?',
        'Reason why not doing the Pre-test',
        'Reason why not doing the Post-test',
        'Student outreached?',
        'owner',
        'modified_by',
    ]

    for col_num in range(len(columns)):
        ws.write(row_num_student, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    qs = queryset_students.extra(select={

        'pre_test_attended_arabic': "pre_test->>'BLN_ASSESSMENT/attended_arabic'",
        'pre_test_modality_arabic': "pre_test->>'BLN_ASSESSMENT/modality_arabic'",
        'pre_test_arabic': "pre_test->>'BLN_ASSESSMENT/arabic'",

        'pre_test_attended_english': "pre_test->>'BLN_ASSESSMENT/attended_english'",
        'pre_test_modality_english': "pre_test->>'BLN_ASSESSMENT/modality_english'",
        'pre_test_english': "pre_test->>'BLN_ASSESSMENT/english'",

        'pre_test_attended_psychomotor': "pre_test->>'BLN_ASSESSMENT/attended_psychomotor'",
        'pre_test_modality_psychomotor': "pre_test->>'BLN_ASSESSMENT/modality_psychomotor'",
        'pre_test_psychomotor': "pre_test->>'BLN_ASSESSMENT/psychomotor'",

        'pre_test_attended_artistic': "pre_test->>'BLN_ASSESSMENT/attended_artistic'",
        'pre_test_modality_artistic': "pre_test->>'BLN_ASSESSMENT/modality_artistic'",
        'pre_test_artistic': "pre_test->>'BLN_ASSESSMENT/artistic'",

        'pre_test_attended_math': "pre_test->>'BLN_ASSESSMENT/attended_math'",
        'pre_test_modality_math': "pre_test->>'BLN_ASSESSMENT/modality_math'",
        'pre_test_math': "pre_test->>'BLN_ASSESSMENT/math'",

        'pre_test_attended_social': "pre_test->>'BLN_ASSESSMENT/attended_social'",
        'pre_test_modality_social': "pre_test->>'BLN_ASSESSMENT/modality_social'",
        'pre_test_social_emotional': "pre_test->>'BLN_ASSESSMENT/social_emotional'",

        'post_test_attended_arabic': "post_test->>'BLN_ASSESSMENT/attended_arabic'",
        'post_test_modality_arabic': "post_test->>'BLN_ASSESSMENT/modality_arabic'",
        'post_test_arabic': "post_test->>'BLN_ASSESSMENT/arabic'",

        'post_test_attended_english': "post_test->>'BLN_ASSESSMENT/attended_english'",
        'post_test_modality_english': "post_test->>'BLN_ASSESSMENT/modality_english'",
        'post_test_english': "post_test->>'BLN_ASSESSMENT/english'",

        'post_test_attended_psychomotor': "post_test->>'BLN_ASSESSMENT/attended_psychomotor'",
        'post_test_modality_psychomotor': "post_test->>'BLN_ASSESSMENT/modality_psychomotor'",
        'post_test_psychomotor': "post_test->>'BLN_ASSESSMENT/psychomotor'",

        'post_test_attended_artistic': "post_test->>'BLN_ASSESSMENT/attended_artistic'",
        'post_test_modality_artistic': "post_test->>'BLN_ASSESSMENT/modality_artistic'",
        'post_test_artistic': "post_test->>'BLN_ASSESSMENT/artistic'",

        'post_test_attended_math': "post_test->>'BLN_ASSESSMENT/attended_math'",
        'post_test_modality_math': "post_test->>'BLN_ASSESSMENT/modality_math'",
        'post_test_math': "post_test->>'BLN_ASSESSMENT/math'",

        'post_test_attended_social': "post_test->>'BLN_ASSESSMENT/attended_social'",
        'post_test_modality_social': "post_test->>'BLN_ASSESSMENT/modality_social'",
        'post_test_social_emotional': "post_test->>'BLN_ASSESSMENT/social_emotional'",
    })

    rows = qs.order_by('id').values_list(
        'id',
        'new_registry',
        'partner__name',
        'round__name',
        'governorate__name_en',
        'district__name_en',
        'cadaster__name_en',
        'location',
        'center__name',
        'language',
        'student__address',
        'registration_level',
        'first_attendance_date',
        'student__id_number',
        'student__number',
        'student__first_name',
        'student__father_name',
        'student__last_name',
        'student__mother_fullname',
        'student__sex',
        'student__nationality__name',
        'other_nationality',
        'student__birthday_day',
        'student__birthday_month',
        'student__birthday_year',
        'student__p_code',
        'disability__name_en',
        'education_status',
        'miss_school_date',
        'internal_number',
        'rims_case_number',
        'source_of_identification',
        'source_of_identification_specify',
        'source_of_transportation',
        'hh_educational_level__name',
        'father_educational_level__name',
        'phone_number',
        'phone_number_confirm',
        'phone_owner',
        'second_phone_number',
        'second_phone_number_confirm',
        'second_phone_owner',
        'main_caregiver',
        'main_caregiver_nationality__name',
        'other_caregiver_relationship',
        'main_caregiver_nationality_other',
        'caretaker_first_name',
        'caretaker_middle_name',
        'caretaker_last_name',
        'caretaker_mother_name',
        'caretaker_birthday_year',
        'caretaker_birthday_month',
        'caretaker_birthday_day',
        'id_type',
        'case_number',
        'case_number_confirm',
        'parent_individual_case_number',
        'parent_individual_case_number_confirm',
        'individual_case_number',
        'individual_case_number_confirm',
        'recorded_number',
        'recorded_number_confirm',
        'parent_national_number',
        'parent_national_number_confirm',
        'national_number',
        'national_number_confirm',
        'parent_syrian_national_number',
        'parent_syrian_national_number_confirm',
        'syrian_national_number',
        'syrian_national_number_confirm',
        'parent_sop_national_number',
        'parent_sop_national_number_confirm',
        'sop_national_number',
        'sop_national_number_confirm',
        'parent_other_number',
        'parent_other_number_confirm',
        'other_number',
        'other_number_confirm',
        'student__family_status',
        'student__have_children',
        'student_number_children',
        'have_labour_single_selection',
        'labours_single_selection',
        'labours_other_specify',
        'labour_hours',
        'labour_weekly_income',
        'participation',
        'barriers_single',
        'barriers_other',
        'test_done',
        'round_complete',
        'basic_stationery',
        'pss_kit',
        'learning_result',
        'learning_result_other',
        'akelius_program',
        'cp_referral',
        'referal_wash',
        'referal_health',
        'referal_other',
        'referal_other_specify',
        'child_received_books',
        'child_received_printout',
        'child_received_internet',
        'phone_call_number',
        'phone_call_follow_up_result',
        'house_visit_number',
        'house_visit_follow_up_result',
        'family_visit_number',
        'family_visit_follow_up_result',
        'parent_attended_visits',
        'pss_session_attended',
        'pss_session_number',
        'pss_session_modality',
        'pss_parent_attended',
        'pss_parent_attended_other',
        'covid_session_attended',
        'covid_session_number',
        'covid_session_modality',
        'covid_parent_attended',
        'covid_parent_attended_other',
        'followup_session_attended',
        'followup_session_number',
        'followup_session_modality',
        'followup_parent_attended',
        'followup_parent_attended_other',
        'pre_test_attended_arabic',
        'pre_test_modality_arabic',
        'pre_test_arabic',
        'pre_test_attended_english',
        'pre_test_modality_english',
        'pre_test_english',
        'pre_test_attended_psychomotor',
        'pre_test_modality_psychomotor',
        'pre_test_psychomotor',
        'pre_test_attended_artistic',
        'pre_test_modality_artistic',
        'pre_test_artistic',
        'pre_test_attended_math',
        'pre_test_modality_math',
        'pre_test_math',
        'pre_test_attended_social',
        'pre_test_modality_social',
        'pre_test_social_emotional',
        'post_test_attended_arabic',
        'post_test_modality_arabic',
        'post_test_arabic',
        'post_test_attended_english',
        'post_test_modality_english',
        'post_test_english',
        'post_test_attended_psychomotor',
        'post_test_modality_psychomotor',
        'post_test_psychomotor',
        'post_test_attended_artistic',
        'post_test_modality_artistic',
        'post_test_artistic',
        'post_test_attended_math',
        'post_test_modality_math',
        'post_test_math',
        'post_test_attended_social',
        'post_test_modality_social',
        'post_test_social_emotional',
        'remote_learning',
        'remote_learning_reasons_not_engaged',
        'reasons_not_engaged_other',
        'reliable_internet',
        'gender_participate',
        'gender_participate_explain',
        'remote_learning_engagement',
        'meet_learning_outcomes',
        'parent_learning_support_rate',
        'covid_message',
        'covid_message_how_often',
        'covid_parents_message',
        'covid_parents_message_how_often',
        'follow_up_done',
        'follow_up_done_with_who',
        'unsuccessful_pretest_reason',
        'unsuccessful_posttest_reason',
        'student_outreached',
        'owner__username',
        'modified_by__username',
        # 'created',
        # 'modified',
    )
    for row in rows:
        row_num_student += 1
        for col_num in range(len(row)):
            ws.write(row_num_student, col_num, row[col_num], font_style)

    # FC
    wsFC = wb_student.add_sheet('FC')
    row_num_fc = 0
    row_num_header_fc = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns_fc_student = [
        'enrollment id',
        'Student first name',
        'Student father name',
        'Student last name',
    ]
    for col_num in range(len(columns_fc_student)):
        wsFC.write(row_num_header_fc, col_num, columns_fc_student[col_num], font_style)

    font_style = xlwt.XFStyle()

    columns_fc = [
        'fc type',
        'facilitator name',
        'subject taught',
        'date of monitoring',
        'targeted competencies',
        'activities reported',
        'activities reported other',
        'share expectations',
        'share expectations no reason',
        'share expectations other reason',
        'materials needed available',
        'attend lesson',
        'child interact teacher',
        'child interact friends',
        'child clear responses',
        'child ask questions',
        'child acquire competency',
        'child show improvement',
        'child expected work independently',
        'work independently evaluation',
        'complete printed package',
        'sessions participated',
        'not participating reason',
        'E recharge card provided',
        'action to taken',
        'action to taken specify',
        'child needs pss',
        'child cant access resources',
        'homework after lesson',
        'parents supporting student',
        'completed tasks',
        'meet objectives',
        'meet objectives verified',
        'objectives verified specify',
        'additional notes'
    ]
    fc_type = [
        'pre-arabic',
        'post-arabic',
        'pre-math',
        'post-math',
        'pre-language',
        'post-language'
    ]
    for subject_type in range(len(fc_type)):
        for col_num in range(len(columns_fc)):
            wsFC.write(row_num_header_fc, col_num + len(columns_fc_student) + (len(columns_fc) * (subject_type))
                       , columns_fc[col_num], font_style)

    font_style = xlwt.XFStyle()

    rows_fc = queryset_fc.values_list(
        'enrollment_id',
        'enrollment__student__first_name',
        'enrollment__student__father_name',
        'enrollment__student__last_name',
        'fc_type',
        'facilitator_name',
        'subject_taught',
        'date_of_monitoring',
        'targeted_competencies',
        'activities_reported',
        'activities_reported_other',
        'share_expectations',
        'share_expectations_no_reason',
        'share_expectations_other_reason',
        'materials_needed_available',
        'attend_lesson',
        'child_interact_teacher',
        'child_interact_friends',
        'child_clear_responses',
        'child_ask_questions',
        'child_acquire_competency',
        'child_show_improvement',
        'child_expected_work_independently',
        'work_independently_evaluation',
        'complete_printed_package',
        'sessions_participated',
        'not_participating_reason',
        'e_recharge_card_provided',
        'action_to_taken',
        'action_to_taken_specify',
        'child_needs_pss',
        'child_cant_access_resources',
        'homework_after_lesson',
        'parents_supporting_student',
        'completed_tasks',
        'meet_objectives',
        'meet_objectives_verified',
        'objectives_verified_specify',
        'additional_notes'

    )
    enrollment_id = 0

    for row in rows_fc:
        row_num_fc += 0
        enrol_id = row[0]

        if enrollment_id != enrol_id:
            row_num_fc += 1
            enrollment_id = enrol_id
            for col_num in range(len(columns_fc_student)):
                wsFC.write(row_num_fc, col_num, row[col_num], font_style)

            row_type = row[4]
            subject_type_index = fc_type.index(row_type)

            for col_num in range(len(columns_fc)):
                wsFC.write(row_num_fc, col_num + len(columns_fc_student) + (len(columns_fc) * (subject_type_index))
                           , row[col_num + len(columns_fc_student)], font_style)

    wb_student.save(buffer)

    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    response = FileResponse(buffer, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="BLN.xls"'

    return response


class MemorySavingQuerysetIterator(object):

    def __init__(self, queryset, max_obj_num=1000):
        self._base_queryset = queryset
        self._generator = self._setup()
        self.max_obj_num = max_obj_num

    def _setup(self):
        for i in xrange(0, self._base_queryset.count(), self.max_obj_num):
            # By making a copy of the queryset and using that to actually access
            # the objects we ensure that there are only `max_obj_num` objects in
            # memory at any given time
            smaller_queryset = copy.deepcopy(self._base_queryset)[i:i + self.max_obj_num]
            # logger.debug('Grabbing next %s objects from DB' % self.max_obj_num)
            for obj in smaller_queryset.iterator():
                yield obj

    def __iter__(self):
        return self

    def next(self):
        return self._generator.next()


def get_outreach_child(outreach_id):
    initial = {}
    instance = OutreachChild.objects.get(id=outreach_id)
    initial['child_outreach'] = instance.id
    initial['student_first_name'] = instance.first_name
    initial['student_father_name'] = instance.outreach_caregiver.father_name
    initial['student_last_name'] = instance.outreach_caregiver.last_name
    initial['student_mother_fullname'] = instance.outreach_caregiver.mother_full_name
    initial['student_birthday_year'] = instance.birthday_year
    initial['student_birthday_month'] = instance.birthday_month
    initial['student_birthday_day'] = instance.birthday_day
    initial['student_sex'] = instance.gender
    nationality = instance.nationality

   # TODO check nationality'palestinian'
    # 3: for "فلسطينية  - من سوريا"
    # 4: for "فلسطنية - من لبنان"
    if nationality == 'syrian':
        initial['student_nationality'] = 1
    elif nationality == 'lebanese':
        initial['student_nationality'] = 5
    # elif nationality == 'palestinian':
    #     initial['student_nationality'] = 4
    elif nationality == 'iraqi':
        initial['student_nationality'] = 2
    elif nationality == 'stateless':
        initial['student_nationality'] = 7
    elif nationality == 'other':
        initial['student_nationality'] = 6
    initial['other_nationality'] = instance.nationality_other
    initial['student_address'] = instance.outreach_caregiver.address

    disability = instance.disability

    # 1	"No"
    if disability == 'no' or disability == 'No':
        initial['disability'] = 1
    # 2	"Other Difficulties"
    elif disability == 'Other':
        initial['disability'] = 2
    # 3	"Difficulty hearing"
    elif disability == 'Difficulty hearing' or disability == 'difficulty_hearing':
        initial['disability'] = 3
    # 4	"Difficulty walking or moving hands"
    elif disability == 'difficulty_walking_or_moving_hands' or disability == 'Difficulty walking or moving hands':
        initial['disability'] = 4
    # 5	"Difficulty Speaking"
    elif disability == 'Difficulty Speaking' or disability == 'difficulty_speaking':
        initial['disability'] = 5
    # 6	"Difficulty seeing"
    elif disability == 'Difficulty seeing' or 'difficulty_seeing':
        initial['disability'] = 6
    # 7	"Difficulty with Self-Care"
    # 8	"Learning Difficulties"
    elif disability == 'learning_difficulties' or disability == 'Learning Difficulties':
        initial['disability'] = 8
    # 9	"Difficulty interacting with others"
    elif disability == 'difficulty_interacting_with_others' or disability == 'Difficulty interacting with others':
        initial['disability'] = 9
    # 10	"Intellectual Disability"
    elif disability == 'intellectual_disability' or disability == 'Intellectual Disability':
        initial['disability'] = 10
    initial['disability_other'] = instance.disability_other

    family_status = instance.family_status.capitalize()
    if family_status == 'Widow':
        initial['student_family_status'] = 'widower'
    elif family_status == 'Separated':
        initial['student_family_status'] = 'divorced'
    else:
        initial['student_family_status'] = family_status


    # TODO check nationality'palestinian'
    # 3: for "فلسطينية  - من سوريا"
    # 4: for "فلسطنية - من لبنان"
    main_caregiver_nationality = instance.outreach_caregiver.caregiver_nationality
    if main_caregiver_nationality == 'syrian':
        initial['main_caregiver_nationality'] = 1
    elif main_caregiver_nationality == 'lebanese':
    #     initial['main_caregiver_nationality'] = 5
    # elif main_caregiver_nationality == 'palestinian':
        initial['main_caregiver_nationality'] = 4
    elif main_caregiver_nationality == 'iraqi':
        initial['main_caregiver_nationality'] = 2
    elif main_caregiver_nationality == 'stateless':
        initial['main_caregiver_nationality'] = 7
    elif main_caregiver_nationality == 'other':
        initial['main_caregiver_nationality'] = 6
    initial['main_caregiver_nationality_other'] = instance.outreach_caregiver.caregiver_nationality_other

    initial['have_labour_single_selection'] = instance.working_status
    if instance.working_status == 'yes':
        initial['have_labour_single_selection'] = 'Yes - All day'

        labour_type = instance.work_type
        if labour_type == 'manufacturing_producing':
            initial['labours_single_selection'] = 'manufacturing'
        elif labour_type == 'garage_mechanics_workshop':
            initial['labours_single_selection'] = ''
        elif labour_type == 'construction_site':
            initial['labours_single_selection'] = 'building'
        elif labour_type == 'shop_restaurant_bakery_barber':
            initial['labours_single_selection'] = 'retail_store'
        elif labour_type == 'street_connected_work__begging__vending_':
            initial['labours_single_selection'] = 'begging'
        elif labour_type == 'agriculture_animal_herding':
            initial['labours_single_selection'] = 'agriculture'
        elif labour_type == 'others':
            initial['labours_single_selection'] = 'other_many_other'
        else:
            initial['labours_single_selection'] = ''

        initial['labours_other_specify'] = instance.work_type_other

    initial['phone_number'] = instance.outreach_caregiver.primary_phone
    initial['phone_number_confirm'] = instance.outreach_caregiver.primary_phone
    initial['second_phone_number'] = instance.outreach_caregiver.secondary_phone
    initial['second_phone_number_confirm'] = instance.outreach_caregiver.secondary_phone

    initial['caretaker_first_name'] = instance.outreach_caregiver.caregiver_first_name
    initial['caretaker_last_name'] = instance.outreach_caregiver.caregiver_last_name
    initial['caretaker_middle_name'] = instance.outreach_caregiver.caregiver_father_name
    initial['caretaker_mother_name'] = instance.outreach_caregiver.caregiver_mother_name

    main_caregiver = instance.outreach_caregiver.main_caregiver
    if main_caregiver == u'الاب':
        initial['main_caregiver'] = 'father'
        initial['caretaker_first_name'] = instance.outreach_caregiver.father_name
        initial['caretaker_last_name'] = instance.outreach_caregiver.last_name
    else:
        if main_caregiver == u'الام':
            initial['main_caregiver'] = 'mother'
        elif main_caregiver == u'اخر':
            initial['main_caregiver'] = 'other'
    dob_string = instance.outreach_caregiver.caregiver_dob
    if dob_string:
        try:
            dob = datetime.strptime(dob_string, '%Y-%m-%d')
            initial['caretaker_dob'] = dob
            initial['caretaker_birthday_year'] = dob.year
            initial['caretaker_birthday_month'] = dob.month
            initial['caretaker_birthday_day'] = dob.day
        except ValueError:
            pass

    id_type = instance.outreach_caregiver.id_type
    if id_type == 'unhcr_registered' or id_type == 'UNHCR registered':
        initial['id_type'] = 'UNHCR Registered'
        initial['case_number'] = instance.outreach_caregiver.unhcr_case_number
        initial['case_number_confirm'] = instance.outreach_caregiver.unhcr_case_number
        initial['parent_individual_case_number'] = instance.outreach_caregiver.caregiver_unhcr_id
        initial['parent_individual_case_number_confirm'] = instance.outreach_caregiver.caregiver_unhcr_id
        initial['individual_case_number'] = instance.child_unhcr_number
        initial['individual_case_number_confirm'] = instance.child_unhcr_number
    elif id_type == 'unhcr_recorded' or id_type == 'UNHCR recorded':
        initial['id_type'] = 'UNHCR Recorded'
        initial['recorded_number'] = instance.outreach_caregiver.unhcr_barcode
        initial['recorded_number_confirm'] = instance.outreach_caregiver.unhcr_barcode
    elif id_type == 'syrian_id' or id_type == 'Syrian ID':
        initial['id_type'] = 'Syrian national ID'
        initial['parent_syrian_national_number'] = instance.outreach_caregiver.caregiver_personal_id
        initial['parent_syrian_national_number_confirm'] = instance.outreach_caregiver.caregiver_personal_id
        initial['syrian_national_number'] = instance.child_personal_id
        initial['syrian_national_number_confirm'] = instance.child_personal_id
    elif id_type == 'palestinian_id' or id_type == 'Palestinian ID':
        initial['id_type'] = 'Palestinian national ID'
        initial['sop_parent_national_number'] = instance.outreach_caregiver.caregiver_personal_id
        initial['sop_parent_national_number_confirm'] = instance.outreach_caregiver.caregiver_personal_id
        initial['sop_national_number'] = instance.child_personal_id
        initial['sop_national_number_confirm'] = instance.child_personal_id
    elif id_type == 'lebanese_id' or id_type == 'Lebanese ID':
        initial['id_type'] = 'Lebanese national ID'
        initial['parent_national_number'] = instance.outreach_caregiver.caregiver_personal_id
        initial['parent_national_number_confirm'] = instance.outreach_caregiver.caregiver_personal_id
        initial['national_number'] = instance.child_personal_id
        initial['national_number_confirm'] = instance.child_personal_id
    elif id_type == 'No_papers' or id_type == 'No papers':
        initial['id_type'] = 'Child have no ID'
    elif id_type == 'other_nationality_id' or id_type == 'Other Nationality ID':
        initial['id_type'] = 'Other nationality'
        initial['parent_other_number'] = instance.outreach_caregiver.caregiver_personal_id
        initial['parent_other_number_confirm'] = instance.outreach_caregiver.caregiver_personal_id





    # TO DO: Awaiting feedback
    # education_status = instance.education_status
    # if education_status == 'never_been_engaged_in_any_type_of_learni' or education_status == 'Never been engaged in any type of learning':
    #     initial['education_status'] = 'No Registered in any school before'
    # # elif education_status == '':
    # #     initial['education_status'] = 'Was registered in BLN program'
    # # elif education_status == 'Previously_enrolled_in_Non-formal_educat' \
    # #     or education_status == 'Previously enrolled in formal education has been out of school for 1 to 2 years'\
    # #     or education_status == 'Previously enrolled in Non-formal education has been out of school for 1 to 2 years'\
    # #     or education_status == 'Previously enrolled in Non-formal education has been out of school for more than 2 years'\
    # #     or education_status == 'was_enrolled_in_non_formal_education_but':
    # #     initial['education_status'] = 'Was registered in formal school and didnt continue'
    # elif education_status == 'was_enrolled_in_formal_education_but_did' \
    #     or education_status == 'Previously enrolled in formal education has been out of school for 1 year'\
    #     or education_status == 'Previously enrolled in formal education has been out of school for more than 2 years'\
    #     or education_status == 'previously_enrolled_in_formal_'\
    #     or education_status == 'Previously_enrolled_in_formal_education_':
    #     initial['labour_type'] = 'Was registered in formal school and didnt continue'
    # else:
    #     initial['education_status'] = ''

    return initial


def load_child_attendance(round_id, attendance_date, school_id, registration_level):
    from datetime import datetime

    result = []
    try:
        if attendance_date is not None:
            attendance_date = datetime.strptime(attendance_date, '%m/%d/%Y')
            attendance = CLMAttendance.objects.filter(
                round_id=round_id,
                attendance_date=attendance_date,
                school_id=school_id,
                registration_level=registration_level
            ).last()

        if attendance:
            attendances = CLMAttendanceStudent.objects.filter(attendance_day=attendance, registration__deleted=False).select_related('student')
            for att in attendances:
                registration = att.registration if att.registration else None
                student = att.student if att.student else None

                result.append({
                    'registration_id': registration.id if registration else None,
                    'child_id': student.id if student else None,
                    'child_fullname': student.full_name if student else None,
                    'child_mother_fullname': student.mother_fullname if student else None,
                    'child_birthday': student.birthday if student else None,
                    'child_nationality': student.nationality.name if student and student.nationality else None,
                    'attended': att.attended,
                    'absence_reason': att.absence_reason,
                    'absence_reason_other': att.absence_reason_other
                })
        else:
            registrations = Bridging.objects.filter(
                round=round_id,
                school=school_id,
                registration_level=registration_level,
                deleted=False
            ).exclude(
                learning_result='dropout', dropout_date__lte=attendance_date
            ).select_related('student')

            for reg in registrations:
                result.append({
                    'registration_id': reg.id,
                    'child_id': reg.student.id,
                    'child_fullname': reg.student.full_name,
                    'child_mother_fullname': reg.student.mother_fullname,
                    'child_birthday': reg.student.birthday,
                    'child_nationality': reg.student.nationality.name,
                    'attended': 'Yes',
                    'absence_reason': '',
                    'absence_reason_other': ''
                })

        return result

    except Exception as ex:
        return []


def create_attendance(data):
    from datetime import datetime
    attendance_date = datetime.strptime(data["attendance_date"], '%m/%d/%Y')
    day_off = data["attendance_day_off"]
    close_reason = data["close_reason"]
    round_id = data["round_id"]
    school_id = data["school_id"]
    registration_level = data["registration_level"]
    children_attendance = data["children_attendance"]

    try:
        with transaction.atomic():  # Ensure atomic transactions
            attendance, created = CLMAttendance.objects.get_or_create(
                round_id=round_id,
                attendance_date=attendance_date,
                school_id=school_id,
                registration_level=registration_level
            )
            attendance.day_off = day_off
            attendance.close_reason = close_reason
            attendance.save()

            student_ids = [child['child_id'] for child in children_attendance]
            registrations = Bridging.objects.filter(student_id__in=student_ids, round_id=round_id)

            for child in children_attendance:
                student_id = child['child_id']
                registration_id = child['registration_id']
                registration = registrations.filter(student_id=student_id).first()

                # Use update_or_create to minimize database hits
                attendance_child, created = CLMAttendanceStudent.objects.update_or_create(
                    attendance_day=attendance,
                    student_id=student_id,
                    registration_id=registration_id,
                    defaults={
                        'attended': child['attended'],
                        'absence_reason': child['absence_reason'],
                        'absence_reason_other': child['absence_reason_other'],
                    }
                )

                update_total_attendance(registration_id, student_id, round_id, school_id, registration.student.first_name,
                                           registration.student.father_name, registration.student.last_name,)

                if child['attended'] == 'No':
                    update_consecutive_absence(registration_id, student_id, registration.student.first_name,
                                               registration.student.father_name, registration.student.last_name,
                                               attendance_date, round_id, school_id, registration)

        return True

    except Exception as ex:
        return False


def update_child_attendance(registration_id, education_program, old_class_section, new_class_section):

    child_attendances = CLMAttendanceStudent.objects.filter(
        registration_id=registration_id,
        attendance_day__education_program=education_program,
        attendance_day__class_section=old_class_section
    )

    try:
        if child_attendances:
            for ca in child_attendances:
                center_id = ca.attendance_day.center.id
                attendance_date = ca.attendance_day.attendance_date

                # Search if attendance for the new class exists and move the child attendance to it
                new_attendance = CLMAttendance.objects.filter(
                    center_id=center_id,
                    attendance_date=attendance_date,
                    education_program=education_program,
                    class_section=new_class_section
                ).last()

                attendance_id = ca.attendance_day.id

                # Count the number of other attendances for the same day
                other_children_count = CLMAttendanceStudent.objects.filter(attendance_day=ca.attendance_day).exclude(id=ca.id).count()

                if new_attendance:
                    ca.attendance_day = new_attendance
                    ca.save()
                else:
                    ca.delete()

                if other_children_count == 0:
                    try:
                        old_attendance = CLMAttendance.objects.get(id=attendance_id)

                        # Delete the unique old_attendance instance
                        old_attendance.delete()

                    except CLMAttendance.DoesNotExist:
                        pass
        return True

    except Exception as ex:
        return False


def update_total_attendance(registration_id, student_id, round_id, school_id, first_name,father_name, last_name):

    total_absence_days = CLMAttendanceStudent.objects.filter(
        attended='No',
        attendance_day__round_id=round_id,
        student_id=student_id
    ).count()
    total_attendance_days = CLMAttendanceStudent.objects.filter(
        attended='Yes',
        attendance_day__round_id=round_id,
        student_id=student_id
    ).count()

    # Update or create the student's total attendance record
    student_attendance, created = CLMStudentTotalAttendance.objects.get_or_create(
        round_id=round_id,
        school_id=school_id,
        registration_id=registration_id,
        defaults={
            'student_first_name': first_name,
            'student_last_name': last_name,
            'student_father_name': father_name,
            'student_id': student_id
        }
    )

    student_attendance.student_first_name = first_name
    student_attendance.student_last_name = last_name
    student_attendance.student_father_name = father_name
    student_attendance.total_attendance_days = total_attendance_days
    student_attendance.total_absence_days = total_absence_days
    student_attendance.save()


def update_consecutive_absence(registration_id, student_id, first_name, father_name, last_name, attendance_date,
                               round_id, school_id, registration):

    working_day_names = School.objects.filter(id=school_id).values_list('working_days', flat=True).first()
    working_days_set = set(working_day_names)

    def get_previous_working_day(current_date):
        previous_day = current_date - timedelta(days=1)
        while previous_day.strftime('%A') not in working_days_set:
            previous_day -= timedelta(days=1)
        return previous_day

    def get_next_working_day(current_date):
        next_day = current_date + timedelta(days=1)
        while next_day.strftime('%A') not in working_days_set:
            next_day += timedelta(days=1)
        return next_day


    previous_absence = CLMStudentAbsences.objects.filter(
        Q(absence_starting_date=get_next_working_day(attendance_date)) |
        Q(absence_ending_date=get_previous_working_day(attendance_date)) |
        Q(absence_starting_date=attendance_date) |
        Q(absence_ending_date=attendance_date),
        registration_id=registration_id,
        round_id=round_id
    ).last()

    if previous_absence:
        attendance_day = attendance_date.date() if isinstance(attendance_date, datetime) else attendance_date
        previous_start_day = previous_absence.absence_starting_date.date() if isinstance(previous_absence.absence_starting_date, datetime) else previous_absence.absence_starting_date
        previous_end_day = previous_absence.absence_ending_date.date() if isinstance(previous_absence.absence_ending_date, datetime) else previous_absence.absence_ending_date

        previous_working_day = get_previous_working_day(previous_start_day)
        if attendance_day == previous_working_day:
            previous_absence.absence_starting_date = attendance_day
            absence_dates_list = previous_absence.absence_dates or []
            absence_dates_list.append(attendance_day.strftime('%Y-%m-%d'))
            previous_absence.absence_dates = absence_dates_list

        next_working_day = get_next_working_day(previous_end_day)
        if attendance_day == next_working_day:
            previous_absence.absence_ending_date = attendance_day
            absence_dates_list = previous_absence.absence_dates or []
            absence_dates_list.append(attendance_day.strftime('%Y-%m-%d'))
            previous_absence.absence_dates = absence_dates_list

        current_day = previous_absence.absence_starting_date
        consecutive_days = 0
        while current_day <= previous_absence.absence_ending_date:
            if current_day.strftime('%A') in working_days_set:
                consecutive_days += 1
            current_day += timedelta(days=1)

        previous_absence.consecutive_absence_days = consecutive_days

        previous_absence.school_id = school_id
        previous_absence.school_name = registration.school.name
        previous_absence.registation_level = registration.registration_level
        previous_absence.student_first_name = first_name
        previous_absence.student_father_name = father_name
        previous_absence.student_last_name = last_name
        previous_absence.save()
    else:
        absence = CLMStudentAbsences(
            partner_id=registration.partner_id,
            registration_id=registration_id,
            student_id=student_id,
            round_id=round_id,
            school_id=school_id,
            school_name=registration.school.name,
            registation_level=registration.registration_level,
            student_first_name=first_name,
            student_father_name=father_name,
            student_last_name=last_name,
            absence_starting_date=attendance_date,
            absence_ending_date=attendance_date,
            absence_dates=[attendance_date.strftime('%Y-%m-%d')],
            consecutive_absence_days=1,
        )
        absence.save()
