# -*- coding: utf-8 -*-
from __future__ import absolute_import, unicode_literals

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponseForbidden, HttpResponse, JsonResponse

from rest_framework import status
from rest_framework import viewsets, mixins, permissions
from django_filters.views import FilterView
from django_tables2 import RequestConfig, SingleTableView
from django_tables2.export.views import ExportMixin
from django.views import View
from django.shortcuts import render, redirect, get_object_or_404
from openpyxl import load_workbook
from django.core.files.base import ContentFile
import csv
import io

from student_registration.adolescent.models import Adolescent
from student_registration.students.models import Nationality, IDType
from student_registration.clm.models import Disability, EducationalLevel
from student_registration.locations.models import Location
from student_registration.youth.models import Registration


from .exporter import export_full_data
from .models import Notification, Exporter, AdolescentUpload
from .serializers import NotificationSerializer, ExporterSerializer
from .filters import ExporterFilter
from .tables import BootstrapTable, ExporterTable
from collections import OrderedDict


def generate_child_unique_id(request):
    from student_registration.backends.threads import generate_child_unique_id

    generate_child_unique_id()
    return HttpResponse("records saved successfully")


def generate_all_child_unique_id(request):
    from student_registration.backends.threads import generate_all_child_unique_id

    generate_all_child_unique_id()
    return HttpResponse("records saved successfully")


def generate_child_cash_programme(request):
    from student_registration.backends.threads import generate_child_programmes

    generate_child_programmes()
    return HttpResponse("records saved successfully")


def generate_student_unique_id(request):
    from student_registration.backends.threads import generate_student_unique_id

    from student_registration.students.models import Student

    generate_student_unique_id()
    return HttpResponse("records saved successfully")


def generate_all_teacher_unique_id(request):
    from student_registration.backends.threads import generate_all_teacher_unique_id

    generate_all_teacher_unique_id()
    return HttpResponse("records saved successfully")



class NotificationViewSet(mixins.UpdateModelMixin,
                          viewsets.GenericViewSet):

    model = Notification
    queryset = Notification.objects.all()
    serializer_class = NotificationSerializer
    permission_classes = (permissions.IsAuthenticated,)

    # def update(self, request, *args, **kwargs):
    #     if 'pk' not in kwargs:
    #         return super(NotificationViewSet, self).update(request)
    #     instance = self.model.objects.get(id=kwargs['pk'])
    #     print(request)
    #     instance.status = True
    #     instance.save()
    #     return JsonResponse({'status': status.HTTP_200_OK, 'data': instance.id})


class ExporterListView(LoginRequiredMixin,
                       FilterView,
                       ExportMixin,
                       SingleTableView,
                       RequestConfig):

    table_class = ExporterTable
    model = Exporter
    template_name = 'backends/files.html'
    table = BootstrapTable(Exporter.objects.all(), order_by='-id')

    filterset_class = ExporterFilter

    def get_queryset(self):
        return Exporter.objects.filter(exported_by=self.request.user)


class ExporterViewSet(LoginRequiredMixin,
                      mixins.ListModelMixin,
                      viewsets.GenericViewSet,):

    model = Exporter
    queryset = Exporter.objects.all()
    serializer_class = ExporterSerializer
    permission_classes = (permissions.IsAuthenticated,)

    def handle_no_permission(self, request):
        return HttpResponseForbidden()

    def list(self, request, *args, **kwargs):
        if self.request.GET.get('report', None):
            #  todo raise a exception if the partner
            data = {
                'report': self.request.GET.get('report'),
                'user': self.request.user.id,
                'partner': self.request.user.partner_id
            }
            export_full_data(data)
        return JsonResponse({'status': status.HTTP_200_OK})


class AdolescentUploadView(LoginRequiredMixin, View):
    template_name = 'backends/adolescent_upload.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return render(request, self.template_name, {'error': 'No file selected'})

        upload = AdolescentUpload.objects.create(file=excel_file, uploaded_by=request.user)
        return redirect('backends:adolescent_upload_confirm', pk=upload.pk)


class AdolescentUploadConfirmView(LoginRequiredMixin, View):
    template_name = 'backends/adolescent_upload_confirm.html'
    result_template = 'backends/adolescent_import_result.html'

    mapping = {
        'Adolescent First Name': 'first_name',
        'Adolescent Father Name': 'father_name',
        'Adolescent Last Name': 'last_name',
        'Adolescent Birthday Year': 'birthday_year',
        'Adolescent Birthday Month': 'birthday_month',
        'Adolescent Birthday Day': 'birthday_day',
        'Gender': 'gender',
        'Adolescent Mother Fullname': 'mother_fullname',
        'Adolescent Nationality': 'nationality',
        'Adolescent Nationality Other': 'nationality_other',
        'Governorate': 'governorate',
        'District': 'district',
        'Cadaster': 'cadaster',
        'Adolescent Address': 'address',
        'Special Need': 'disability',
        'Father Educational Level': 'father_educational_level',
        'Mother Educational Level': 'mother_educational_level',
        'First Phone Number': 'first_phone_number',
        'Second Phone Number': 'second_phone_number',
        'Main Caregiver': 'main_caregiver',
        'Main Caregiver Other': 'main_caregiver_other',
        'Caregiver First Name': 'caregiver_first_name',
        'Caregiver Middle Name': 'caregiver_middle_name',
        'Caregiver Last Name': 'caregiver_last_name',
        'Main Caregiver Nationality Name': 'main_caregiver_nationality',
        'Main Caregiver Nationality Other': 'main_caregiver_nationality_other',
        'ID Type': 'id_type',
        'UNHCR Case Number': 'case_number',
        'Cargiver Individual ID': 'parent_individual_case_number',
        'Individual ID of the youth': 'individual_case_number',
        'UNHCR Barcode number (Shifra number)': 'recorded_number',
        'unrwa_number': 'unrwa_number',
        'Syrian National ID number of the Cargiver': 'parent_syrian_national_number',
        'Syrian National ID number of the youth': 'syrian_national_number',
        'Palestinian ID number of the Cargiver': 'parent_sop_national_number',
        'Palestinian ID number of the youth': 'sop_national_number',
        'Lebanese ID number of the Cargiver': 'parent_national_number',
        'Lebanese ID number of the youth': 'national_number',
        'ID number of the Cargiver': 'parent_other_number',
        'ID number of the youth': 'other_number',
    }

    mandatory_fields = [
        'first_name',
        'father_name',
        'last_name',
        'birthday_year',
        'birthday_month',
        'birthday_day',
        'gender',
        'mother_fullname',
        'nationality',
        'governorate',
        'district',
        'cadaster',
        'disability',
        'first_phone_number',
    ]

    def get(self, request, pk, *args, **kwargs):
        upload = get_object_or_404(AdolescentUpload, pk=pk, uploaded_by=request.user)
        data = self.parse_file(upload.file)

        if isinstance(data, dict) and data.get("error"):
            messages.error(request, data["error"])
            return redirect("backends:adolescent_upload")

        preview = data[:5]
        return render(request, self.template_name, {
            'upload': upload,
            'count': len(data),
            'preview': preview,
        })

    def post(self, request, pk, *args, **kwargs):
        upload = get_object_or_404(AdolescentUpload, pk=pk, uploaded_by=request.user)
        data = self.parse_file(upload.file)

        if isinstance(data, dict) and data.get("error"):
            messages.error(request, data["error"])
            return redirect("backends:adolescent_upload")

        imported, not_imported = self.import_data(data, upload, request)
        upload.processed = True
        upload.save()
        return render(request, self.result_template, {
            'imported': imported,
            'failed': len(not_imported),
            'not_imported': not_imported,
            'upload': upload,
        })

    def parse_file(self, uploaded_file):
        from openpyxl import load_workbook
        try:
            wb = load_workbook(filename=uploaded_file, read_only=True)
        except Exception as e:
            raise ValueError("Failed to read Excel file: {}".format(e))

        if not wb.sheetnames:
            raise ValueError("The uploaded Excel file has no sheets. Please check the file content.")

        if 'Registrations' not in wb.sheetnames:
            raise ValueError("Sheet 'Registrations' not found.")

        ws = wb['Registrations']

        header_row = next(ws.iter_rows(min_row=1, max_row=1))
        headers = [cell.value for cell in header_row]

        rows = []
        for row in ws.iter_rows(min_row=2):
            values = [cell.value for cell in row]
            if not any(values):
                continue
            rows.append(OrderedDict(
                (self.mapping.get(headers[i]), values[i])
                for i in range(len(headers))
                if headers[i] in self.mapping
            ))

        return rows

    def import_data(self, data, upload, request):
        from student_registration.students.utils import generate_one_unique_id
        not_imported = []
        imported = 0
        for index, values in enumerate(data, start=2):
            missing = [f for f in self.mandatory_fields if not values.get(f)]
            if missing:
                values['row'] = index
                values['error'] = 'Missing fields: ' + ', '.join(missing)
                not_imported.append(values)
                continue

            nationality = Nationality.objects.filter(name=values.get('nationality')).first()
            gov = Location.objects.filter(name=values.get('governorate')).first()
            dist = Location.objects.filter(name=values.get('district')).first()
            cad = Location.objects.filter(name=values.get('cadaster')).first()
            disability = Disability.objects.filter(name=values.get('disability')).first()

            if not all([nationality, gov, dist, cad, disability]):
                values['row'] = index
                values['error'] = 'Invalid reference'
                not_imported.append(values)
                continue

            father_ed = EducationalLevel.objects.filter(name=values.get('father_educational_level')).first() if values.get('father_educational_level') else None
            mother_ed = EducationalLevel.objects.filter(name=values.get('mother_educational_level')).first() if values.get('mother_educational_level') else None
            caregiver_nat = Nationality.objects.filter(name=values.get('main_caregiver_nationality')).first() if values.get('main_caregiver_nationality') else None
            id_type = IDType.objects.filter(name=values.get('id_type')).first() if values.get('id_type') else None

            try:
                adolescent = Adolescent.objects.create(
                    first_name=values.get('first_name'),
                    father_name=values.get('father_name'),
                    last_name=values.get('last_name'),
                    birthday_year=values.get('birthday_year'),
                    birthday_month=values.get('birthday_month'),
                    birthday_day=values.get('birthday_day'),
                    gender=values.get('gender'),
                    mother_fullname=values.get('mother_fullname'),
                    nationality=nationality,
                    nationality_other=values.get('nationality_other'),
                    governorate=gov,
                    district=dist,
                    cadaster=cad,
                    address=values.get('address'),
                    disability=disability,
                    father_educational_level=father_ed,
                    mother_educational_level=mother_ed,
                    first_phone_number=values.get('first_phone_number'),
                    second_phone_number=values.get('second_phone_number'),
                    main_caregiver=values.get('main_caregiver'),
                    main_caregiver_other=values.get('main_caregiver_other'),
                    caregiver_first_name=values.get('caregiver_first_name'),
                    caregiver_middle_name=values.get('caregiver_middle_name'),
                    caregiver_last_name=values.get('caregiver_last_name'),
                    main_caregiver_nationality=caregiver_nat,
                    main_caregiver_nationality_other=values.get('main_caregiver_nationality_other'),
                    id_type=id_type,
                    case_number=values.get('case_number'),
                    parent_individual_case_number=values.get('parent_individual_case_number'),
                    individual_case_number=values.get('individual_case_number'),
                    recorded_number=values.get('recorded_number'),
                    unrwa_number=values.get('unrwa_number'),
                    parent_syrian_national_number=values.get('parent_syrian_national_number'),
                    syrian_national_number=values.get('syrian_national_number'),
                    parent_sop_national_number=values.get('parent_sop_national_number'),
                    sop_national_number=values.get('sop_national_number'),
                    parent_national_number=values.get('parent_national_number'),
                    national_number=values.get('national_number'),
                    parent_other_number=values.get('parent_other_number'),
                    other_number=values.get('other_number'),
                )
                if adolescent:
                    adolescent.unicef_id = generate_one_unique_id(
                        str(adolescent.pk),
                        adolescent.first_name,
                        adolescent.father_name,
                        adolescent.last_name,
                        adolescent.mother_fullname,
                        adolescent.birthdate,
                        adolescent.nationality_name_en,
                        adolescent.gender
                    )
                    adolescent.save()

                    Registration.objects.create(
                        adolescent=adolescent,
                        owner=request.user,
                        partner_id=getattr(request.user, 'partner_id', None),
                        center_id=getattr(request.user, 'center_id', None),
                    )
                imported += 1
            except Exception as ex:
                values['row'] = index
                values['error'] = str(ex)
                not_imported.append(values)

        if not_imported:
            csv_buffer = io.StringIO()
            writer = csv.DictWriter(csv_buffer, fieldnames=not_imported[0].keys())
            writer.writeheader()
            writer.writerows(not_imported)
            upload.failed_file.save(
                'failed_{}.csv'.format(upload.pk),
                ContentFile(csv_buffer.getvalue().encode('utf-8'))  # convert str to bytes
            )

        return imported, not_imported


class AdolescentUploadFailedView(LoginRequiredMixin, View):
    def get(self, request, pk, *args, **kwargs):
        upload = get_object_or_404(AdolescentUpload, pk=pk, uploaded_by=request.user)
        if not upload.failed_file:
            return HttpResponse(status=404)
        response = HttpResponse(upload.failed_file, content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=%s' % upload.failed_file.name.split('/')[-1]
        return response

