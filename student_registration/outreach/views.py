# -*- coding: utf-8 -*-
from __future__ import absolute_import, unicode_literals

# from rest_framework.response import Response
import requests
from requests.structures import CaseInsensitiveDict
import json

from django.db.models import Max
from django.views.generic import DetailView, ListView, RedirectView,FormView, TemplateView, UpdateView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required

from django.http import HttpResponse, JsonResponse
from rest_framework import viewsets, mixins, permissions
from openpyxl import Workbook
from .models import HouseHold, Child, OutreachCaregiver, OutreachChild
from .serializers import HouseHoldSerializer, ChildSerializer
import datetime
from django.http import StreamingHttpResponse
from django.db import connection
import math
import io

MAX_CHUNKS = 5


def outreach_import_data(request):
    last_loaded_identifier = OutreachCaregiver.objects.filter(form_id='aLkUps4YnMc43ATvTr9JB3').aggregate(Max('u_id'))['u_id__max']
    if last_loaded_identifier is None:
        last_loaded_identifier = 0
    # last_loaded_identifier = 0
    last_loaded_identifier_str = str(last_loaded_identifier)
    url = "https://eu.kobotoolbox.org/api/v2/assets/aLkUps4YnMc43ATvTr9JB3/data.json?sort=%7B%22_id%22%3A+1%7D&query=%7B%22_id%22%3A+%7B%22%24gt%22%3A+" + last_loaded_identifier_str + "%7D%7D"
    headers = CaseInsensitiveDict()
    headers["Authorization"] = "Token 96d9b5c22e092b684544167a136fba0b62df4c25"
    resp = requests.get(url, headers=headers)
    data = json.loads(resp.text)

    for record in data["results"]:
        try:
            caregiver = OutreachCaregiver()
            caregiver.u_id = record["_id"]
            caregiver.form_id = record["_xform_id_string"]
            record_value(caregiver, "partner_name", record, "partner_name")
            record_value(caregiver, "governorate", record, "Governorate")
            record_value(caregiver, "district", record, "Districts")
            record_value(caregiver, "cadaster", record, "Cadaster")
            record_value(caregiver, "cadaster_other_specify", record, "cadaster_other_specify")
            record_value(caregiver, "address", record, "address")
            record_value(caregiver, "gps", record, "gps")
            record_value(caregiver, "primary_phone", record, "primary_phone")
            record_value(caregiver, "secondary_phone", record, "secondary_phone")
            record_value(caregiver, "father_name", record, "father_name")
            record_value(caregiver, "mother_full_name", record, "mother_full_name")
            record_value(caregiver, "last_name", record, "last_name")
            record_value(caregiver, "main_caregiver", record, "main_caregiver")
            record_value(caregiver, "caregiver_nationality", record, "caregiver_nationality")
            record_value(caregiver, "caregiver_nationality_other", record, "caregiver_nationality_other")
            record_value(caregiver, "caregiver_first_name", record, "caregiver_first_name")
            record_value(caregiver, "caregiver_father_name", record, "caregiver_father_name")
            record_value(caregiver, "caregiver_mother_name", record, "caregiver_mother")
            record_value(caregiver, "caregiver_last_name", record, "caregiver_last_name")
            record_value(caregiver, "caregiver_dob", record, "caregiver_dob")
            record_value(caregiver, "cash_assistance", record, "cash_assistance")
            record_value(caregiver, "id_type", record, "id_type")
            record_value(caregiver, "unhcr_case_number", record, "unhcr_case_number")
            record_value(caregiver, "caregiver_unhcr_id", record, "caregiver_unhcr_id")
            record_value(caregiver, "unhcr_barcode", record, "unhcr_barcode")
            record_value(caregiver, "caregiver_personal_id", record, "caretaker_personal_id")
            record_value(caregiver, "father_education_level", record, "father_education_level")
            record_value(caregiver, "mother_education_level", record, "mother_education_level")
            record_value(caregiver, "other_education_level", record, "other_education_level")
            record_value(caregiver, "number_of_children", record, "DC_count")
            record_value(caregiver, "geolocation", record, "_geolocation")
            record_value(caregiver, "interview_date", record, "_submission_time")
            record_value(caregiver, "submitted_by", record, "_submitted_by")
            record_value(caregiver, "interview_comment", record, "child_notes")
            record_value(caregiver, "submission_status", record, "_status")
            caregiver.save()

            # caregiver_id = caregiver.id
            if "DC" in record:
                try:
                    for student in record["DC"]:
                        st = OutreachChild()
                        st.outreach_caregiver = caregiver
                        record_value(st, "first_name", student, "DC/first_name")
                        record_value(st, "date_of_birth", student, "DC/date_of_birth")
                        if "DC/date_of_birth" in student:
                            dt_string = student["DC/date_of_birth"]
                            try:
                                dt = datetime.datetime.strptime(dt_string, '%Y-%m-%d')
                                st.birthday_year = dt.year
                                st.birthday_month = dt.month
                                st.birthday_day = dt.day
                            except ValueError as e:
                                pass

                        if "DC/gender" in student:
                            if student["DC/gender"] == '_':
                                st.gender = "Male"
                            elif student["DC/gender"] == '__1':
                                st.gender = "Female"
                            else:
                                st.gender = student["DC/gender"]
                        record_value(st, "nationality", student, "DC/nationality")
                        record_value(st, "nationality_other", student, "DC/nationality_other")
                        record_value(st, "child_unhcr_number", student, "DC/child_unhcr_number")
                        record_value(st, "child_personal_id", student, "DC/child_personal_id")
                        record_value(st, "family_status", student, "DC/family_status")
                        record_value(st, "disability", student, "DC/disability")
                        record_value(st, "disability_other", student, "DC/disability_other")
                        record_value(st, "education_status", student, "DC/education_status")
                        record_value(st, "dropout_date", student, "DC/dropout_date")
                        record_value(st, "dropout_reason", student, "DC/dropout_reason")
                        record_value(st, "dropout_reason_other", student, "DC/dropout_reason_other")
                        record_value(st, "working_status", student, "DC/working_status")
                        record_value(st, "work_type", student, "DC/work_type")
                        record_value(st, "work_type_other", student, "DC/work_type_other")
                        record_value(st, "child_referral", student, "DC/child_referral")
                        record_value(st, "child_notes", student, "DC/child_notes")
                        st.save()
                except Exception as student_error:
                    pass

        except Exception as record_error:
            pass
    return HttpResponse("records saved successfully")


def record_value(model,model_field,record, field):
    if field in record:
        model.__setattr__(model_field, record[field])


class HouseHoldViewSet(mixins.RetrieveModelMixin,
                       mixins.ListModelMixin,
                       mixins.CreateModelMixin,
                       mixins.UpdateModelMixin,
                       viewsets.GenericViewSet):

    model = HouseHold
    queryset = HouseHold.objects.all()
    serializer_class = HouseHoldSerializer
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        if self.request.method in ["PATCH", "POST", "PUT"]:
            return self.queryset
        term = self.request.GET.get('term', 0)
        if term:
            qs = self.queryset.filter(barcode_number=term).distinct()
            return qs
        return []


class ChildViewSet(mixins.RetrieveModelMixin,
                   mixins.ListModelMixin,
                   mixins.CreateModelMixin,
                   mixins.UpdateModelMixin,
                   viewsets.GenericViewSet):

    model = Child
    queryset = Child.objects.all()
    serializer_class = ChildSerializer
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        if self.request.method in ["PATCH", "POST", "PUT"]:
            return self.queryset
        term = self.request.GET.get('term', 0)
        if term:
            qs = self.queryset.filter(barcode_subset__contains=term).distinct()
            return qs
        return []


class OutreachPage(LoginRequiredMixin,
                   TemplateView):
    template_name = 'outreach/outreach.html'


@login_required(login_url='/users/login')
def outreach_export_data(request):
    from django.db import connection
    cursor = connection.cursor()
    vw_data = 'SELECT * FROM vw_outreach_data WHERE caregiver_id > 0'

    cursor.execute(vw_data)
    data = cursor.fetchall()

    headers = [col[0] for col in cursor.description]

    workbook = Workbook()
    worksheet_all_data = workbook.create_sheet("Outreach Data")
    worksheet_all_data.append(headers)

    for row in data:
        worksheet_all_data.append(row)

    default_sheet = workbook.get_sheet_by_name('Sheet')
    workbook.remove(default_sheet)

    # Set the appropriate response headers for the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=outreach_data.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response


@login_required(login_url='/users/login')
def outreach_unregistered_export_info(request):
    cursor = connection.cursor()
    cursor.execute('SELECT COUNT(*) FROM vw_outreach_not_registered WHERE caregiver_id > 0')
    total_records = cursor.fetchone()[0]
    return JsonResponse({'total_records': total_records})


@login_required(login_url='/users/login')
def outreach_unregistered_export_data(request, part):
    part = int(part)
    cursor = connection.cursor()
    cursor.execute('SELECT COUNT(*) FROM vw_outreach_not_registered WHERE caregiver_id > 0')
    total_records = cursor.fetchone()[0]
    records_per_chunk = math.ceil(total_records / MAX_CHUNKS)
    offset = (part - 1) * records_per_chunk

    vw_data = 'SELECT * FROM vw_outreach_not_registered WHERE caregiver_id > 0 LIMIT {} OFFSET {}'.format(records_per_chunk, offset)
    cursor.execute(vw_data)
    data = cursor.fetchall()
    headers = [col[0] for col in cursor.description]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Part {}".format(part)
    worksheet.append(headers)
    for row in data:
        worksheet.append(row)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    response = StreamingHttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=outreach_not_registered_data_part_{}.xlsx'.format(part)
    return response
