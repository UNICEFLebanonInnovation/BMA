# -*- coding: utf-8 -*-
from __future__ import absolute_import, unicode_literals
import json
import io
import xlwt
import csv
import logging
import traceback
logging.basicConfig(level=logging.ERROR)
import codecs
import datetime
import zipfile
import os
import uuid
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile

from django.views.generic import DetailView, ListView, RedirectView, UpdateView, CreateView
from django.forms import inlineformset_factory
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q, Sum, Avg, F, Func, When ,OuterRef, Subquery
from django.http import HttpResponse, JsonResponse, FileResponse
from django.contrib import messages

from django_tables2 import MultiTableMixin, RequestConfig, SingleTableView
from django.urls import reverse
from django.shortcuts import redirect
from braces.views import GroupRequiredMixin, SuperuserRequiredMixin
from django.contrib.auth.decorators import login_required
from rest_framework import viewsets, mixins, permissions
from rest_framework.generics import ListAPIView
from rest_framework import status

from django.utils.translation import gettext as _
from django.views.generic.edit import FormView

from student_registration.schools.models import (
    School,
    Section,
    ClassRoom,
    EducationLevel,
)
from student_registration.enrollments.models import (
    Enrollment,
)
from student_registration.alp.models import Outreach, ALPRound
from student_registration.backends.tasks import export_attendance
from student_registration.users.utils import force_default_language
from .utils import find_attendances, fill_attendancedt
# calculate_absentees
from .models import Attendance, Absentee, CLMAttendance, CLMAttendanceStudent, CLMStudentAbsences, CLMStudentTotalAttendance
from student_registration.clm.models import Bridging
from student_registration.schools.models import CLMRound
from student_registration.backends.models import ExportHistory

from .serializers import AttendanceSerializer, AbsenteeSerializer, AttendanceExportSerializer

from .forms import MainAttendanceForm, AttendanceStudentForm, AttendanceAbsenceForm


class AttendanceViewSet(mixins.RetrieveModelMixin,
                        mixins.ListModelMixin,
                        mixins.CreateModelMixin,
                        mixins.UpdateModelMixin,
                        viewsets.GenericViewSet):
    model = Attendance
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        if not self.request.user.is_superuser:
            if self.request.user.school:
                return self.queryset.filter(school_id=self.request.user.school.id)
            else:
                return []

        return self.queryset

    def list(self, request, *args, **kwargs):
        if self.request.GET.get('from_date', None) and self.request.GET.get('to_date', None):
            data = find_attendances(governorate=self.request.GET.get('governorate', None),
                                    student_id=self.request.GET.get('student', None),
                                    from_date=self.request.GET.get('from_date', None),
                                    to_date=self.request.GET.get('to_date', None),
                                    filter_by_status=request.GET.get('status', False)
                                    )
            return JsonResponse(json.dumps(data), safe=False)

        return JsonResponse({'status': status.HTTP_200_OK})

    def create(self, request, *args, **kwargs):
        """
        :return: JSON
        """
        try:
            instance = Attendance.objects.get(school=int(request.POST.get('school')),
                                              attendance_date=request.POST.get('attendance_date'),
                                              school_type=request.POST.get('school_type'))
            return JsonResponse({'status': status.HTTP_200_OK, 'data': instance.id})
        except Attendance.DoesNotExist:
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.instance = serializer.save()
            return JsonResponse({'status': status.HTTP_201_CREATED, 'data': serializer.instance.id})

    def update(self, request, *args, **kwargs):
        if 'pk' not in kwargs:
            return super(AttendanceViewSet, self).update(request)
        instance = self.model.objects.get(id=kwargs['pk'])
        data = {}
        level_section = ''
        if request.data.keys():
            data = json.loads(request.data.keys()[0], "utf-8")
            level_section = data.keys()[0]
        if not instance.students:
            instance.students = data
        elif level_section:
            instance.students[level_section] = data[level_section]
        instance.save()
        if instance.students:
            fill_attendancedt(instance, data[level_section]['students'])
        # calculate_absentees(instance, data[level_section]['students'])
        return JsonResponse({'status': status.HTTP_200_OK, 'data': instance.id})

    def perform_update(self, serializer):
        instance = serializer.save()
        instance.save()

    def partial_update(self, request, *args, **kwargs):
        return super(AttendanceViewSet, self).partial_update(request)


class AbsenteeViewSet(mixins.ListModelMixin,
                      viewsets.GenericViewSet):
    model = Absentee
    queryset = Absentee.objects.all()
    serializer_class = AbsenteeSerializer
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        if self.request.GET.get('days', None):
            return self.queryset.filter(absent_days__lte=self.request.GET.get('days', None), absent_days__gte=5)
        # .exclude(student__student_enrollment__dropout_status=True)
        # .exclude(student__student_enrollment__disabled=True)
        return []

    # def list(self, request, *args, **kwargs):
    #     if self.request.GET.get('days', None):
    #         return self.queryset.filter(absent_days=self.request.GET.get('days', None))
    #     return JsonResponse({'status': status.HTTP_200_OK})


class AttendanceView(LoginRequiredMixin,
                     GroupRequiredMixin,
                     ListView):
    model = Attendance
    template_name = 'attendances/school_day.html'
    group_required = [u"ATTENDANCE"]

    def get_context_data(self, **kwargs):
        force_default_language(self.request)
        level = 0
        section = 0
        school = 0
        levels_by_sections = []
        attendance_students = []
        attendance_status = {}
        students = []
        date_format = '%Y-%m-%d'
        date_format_display = '%A %d/%m/%Y'

        if self.request.user.school:
            school = self.request.user.school

        if not school.academic_year_start:
            messages.warning(self.request, _('Please go to the school profile and enter the academic start date '
                                             'in order to take attendance.'))
            self.template_name = 'error.html'
            return {
            }

        current_date = datetime.datetime.now().strftime(date_format)
        selected_date = self.request.GET.get('date', current_date)
        selected_date_obj = datetime.datetime.strptime(selected_date, date_format).date()
        selected_date_view = datetime.datetime.strptime(selected_date, date_format).strftime(date_format_display)

        try:
            attendance = Attendance.objects.get(
                school_id=school.id,
                attendance_date=selected_date,
                school_type='2nd-shift',
                education_year__current_year=True
            )
        except Attendance.DoesNotExist:
            attendance = ''

        if self.request.GET.get('level', 0):
            level = ClassRoom.objects.get(id=int(self.request.GET.get('level', 0)))
            self.template_name = 'attendances/level_section.html'
        if self.request.GET.get('section', 0):
            section = Section.objects.get(id=int(self.request.GET.get('section', 0)))

        # education_year = EducationYear.objects.get(current_year=True)
        queryset = Enrollment.objects.exclude(last_moved_date__lt=selected_date,
                                              moved=True).filter(school_id=school, education_year__current_year=True,
                                                                 dropout_status=False)
        # queryset = Enrollment.objects.exclude(moved=True).filter(school_id=school, education_year=education_year)
        registrations = queryset.filter(
            classroom__isnull=False,
            section__isnull=False
        ).distinct().values(
            'classroom__name',
            'classroom_id',
            'section__name',
            'section_id'
        ).order_by('classroom_id')

        current_level_section = ''
        disable_attendance = False
        for registry in registrations:
            exam_day = False
            not_attending = False
            school_closed = attendance.close_reason if attendance else False
            validation_date = attendance.validation_date if attendance else ''
            total_attended = 0
            total_absences = 0
            attendance_taken = False
            level_section = '{}-{}'.format(registry['classroom_id'], registry['section_id'])
            attendances = attendance.students[
                level_section] if attendance and attendance.students and level_section in attendance.students else ''
            total = queryset.filter(classroom_id=registry['classroom_id'],
                                    section_id=registry['section_id'],
                                    registration_date__lte=selected_date).count()
            if total == 0:
                continue

            if attendances:
                attendance_taken = True
                total = attendances['total_enrolled']
                total_attended = attendances['total_attended']
                total_absences = attendances['total_absences']
                exam_day = attendances['exam_day'] if 'exam_day' in attendances else False
                not_attending = attendances['not_attending'] if 'not_attending' in attendances else False
                for value in attendances['students']:
                    attendance_status[value['student_id']] = value

            level_by_section = {
                'level_name': registry['classroom__name'],
                'level': registry['classroom_id'],
                'section_name': registry['section__name'],
                'section': registry['section_id'],
                'total': total,
                'total_attended': total_attended,
                'total_absences': total_absences,
                'exam_day': exam_day,
                'not_attending': not_attending,
                'validation_date': validation_date,
                'disable_attendance': disable_attendance,
                'attendance_taken': attendance_taken,
                'school_closed': school_closed
            }

            if level and section and level.id == registry['classroom_id'] \
                and section.id == registry['section_id']:
                current_level_section = level_by_section
                if exam_day or not_attending or (attendance and attendance.validation_date) or school_closed:
                    disable_attendance = True

            levels_by_sections.append(level_by_section)

        if attendance and (attendance.validation_date or attendance.close_reason):
            disable_attendance = True

        if level and section:
            students = queryset.filter(classroom_id=level.id,
                                       section_id=section.id,
                                       registration_date__lte=selected_date,
                                       ).order_by('student__first_name', 'student__father_name', 'student__last_name')
            for line in students:
                student = line.student
                if str(student.id) in attendance_status:
                    student_status = attendance_status[str(student.id)]
                    line.attendance_status = student_status['status'] if 'status' in student_status else True
                    line.absence_reason = student_status['absence_reason'] if 'absence_reason' in student_status else ''
                    attendance_students.append(line)

        base = datetime.datetime.now()
        dates = []
        allowed_dates = []
        if school.attendance_from_beginning:
            start_date = school.academic_year_start
            end_date = datetime.date(base.year, base.month, base.day)
            delta = end_date - start_date
            day_range = delta.days + 1
        else:
            day_range = school.attendance_range if school.attendance_range else Attendance.DEFAULT_ATTENDANCE_RANGE

        for x in range(0, day_range):
            d = base - datetime.timedelta(days=x)
            allowed_dates.append(d.strftime(date_format))
            dates.append({
                'value': d.strftime(date_format),
                'label': d.strftime(date_format_display)
            })

        if selected_date not in allowed_dates:
            messages.warning(self.request,
                             _('This dates is blocked you are not allowed to take attendance for this date.'))
            self.template_name = 'error.html'
            return {
            }

        return {
            'school_type': '2nd-shift',
            'attendance': attendance,
            'disable_attendance': disable_attendance,
            'current_level_section': current_level_section,
            'total': queryset.count(),
            'total_students': students.count() if students else 0,
            'students': students,
            'school': school,
            'level': level,
            'section': section,
            'dates': dates,
            'classrooms': ClassRoom.objects.all(),
            'sections': Section.objects.all(),
            'levels_by_sections': levels_by_sections,
            'selected_date': selected_date,
            'selected_date_view': selected_date_view,
        }


class AbsenteeView(ListAPIView):
    """
    API endpoint for validated absentees
    """
    queryset = Absentee.objects.filter(
        school__location=True
    )
    serializer_class = AbsenteeSerializer
    permission_classes = (permissions.IsAdminUser,)


class AttendanceALPView(LoginRequiredMixin,
                        GroupRequiredMixin,
                        ListView):
    model = Attendance
    template_name = 'attendances/school_day.html'
    group_required = [u"ATTENDANCE"]

    def get_context_data(self, **kwargs):
        force_default_language(self.request)
        level = 0
        section = 0
        school = 0
        levels_by_sections = []
        attendance_students = []
        attendance_status = {}
        students = []
        date_format = '%Y-%m-%d'
        date_format_display = '%A %d/%m/%Y'

        alp_round = ALPRound.objects.get(current_round=True)

        if self.request.user.school:
            school = self.request.user.school

        if not school.academic_year_start:
            messages.warning(self.request,
                             _('Please go to the school profile and enter the academic start date in order to take attendance.'))
            self.template_name = 'error.html'
            return {
            }

        current_date = datetime.datetime.now().strftime(date_format)
        selected_date = self.request.GET.get('date', current_date)
        selected_date_view = datetime.datetime.strptime(selected_date, date_format).strftime(date_format_display)

        try:
            attendance = Attendance.objects.get(
                school_id=school.id,
                attendance_date=selected_date,
                school_type='ALP',
                alp_round__current_round=True
            )
        except Attendance.DoesNotExist:
            attendance = ''

        if self.request.GET.get('level', 0):
            level = EducationLevel.objects.get(id=int(self.request.GET.get('level', 0)))
            self.template_name = 'attendances/level_section.html'
        if self.request.GET.get('section', 0):
            section = Section.objects.get(id=int(self.request.GET.get('section', 0)))

        queryset = Outreach.objects.filter(school_id=school, alp_round__current_round=True)
        registrations = queryset.filter(
            registered_in_level__isnull=False,
            section__isnull=False
        ).distinct().values(
            'registered_in_level__name',
            'registered_in_level_id',
            'section__name',
            'section_id'
        ).order_by('registered_in_level_id')

        current_level_section = ''
        disable_attendance = False
        for registry in registrations:
            exam_day = False
            not_attending = False
            school_closed = attendance.close_reason if attendance else False
            validation_date = attendance.validation_date if attendance else ''
            total_attended = 0
            total_absences = 0
            attendance_taken = False
            level_section = '{}-{}'.format(registry['registered_in_level_id'], registry['section_id'])
            attendances = attendance.students[level_section] if attendance \
                                                                and attendance.students \
                                                                and level_section in attendance.students else ''
            total = queryset.filter(registered_in_level_id=registry['registered_in_level_id'],
                                    section_id=registry['section_id']).count()

            if attendances:
                attendance_taken = True
                total = attendances['total_enrolled']
                total_attended = attendances['total_attended']
                total_absences = attendances['total_absences']
                exam_day = attendances['exam_day'] if 'exam_day' in attendances else False
                not_attending = attendances['not_attending'] if 'not_attending' in attendances else False
                for value in attendances['students']:
                    attendance_status[value['student_id']] = value

            level_by_section = {
                'level_name': registry['registered_in_level__name'],
                'level': registry['registered_in_level_id'],
                'section_name': registry['section__name'],
                'section': registry['section_id'],
                'total': total,
                'total_attended': total_attended,
                'total_absences': total_absences,
                'exam_day': exam_day,
                'not_attending': not_attending,
                'validation_date': validation_date,
                'disable_attendance': disable_attendance,
                'attendance_taken': attendance_taken,
                'school_closed': school_closed
            }

            if level and section and level.id == registry['registered_in_level_id'] and section.id == registry[
                'section_id']:
                current_level_section = level_by_section
                if exam_day or not_attending or (attendance and attendance.validation_date) or school_closed:
                    disable_attendance = True

            levels_by_sections.append(level_by_section)

        if attendance and (attendance.validation_date or attendance.close_reason):
            disable_attendance = True

        if level and section:
            students = queryset.filter(registered_in_level_id=level.id, section_id=section.id,
                                       ).order_by('student__first_name', 'student__father_name', 'student__last_name')
            for line in students:
                student = line.student
                if str(student.id) in attendance_status:
                    student_status = attendance_status[str(student.id)]
                    line.attendance_status = student_status['status'] if 'status' in student_status else True
                    line.absence_reason = student_status['absence_reason'] if 'absence_reason' in student_status else ''
                    attendance_students.append(line)

        base = datetime.datetime.now()
        dates = []
        allowed_dates = []
        if alp_round.round_start_date:
            start_date = alp_round.round_start_date
            end_date = datetime.date(base.year, base.month, base.day)
            delta = end_date - start_date
            day_range = delta.days + 1
        else:
            day_range = school.attendance_range if school.attendance_range else Attendance.DEFAULT_ATTENDANCE_RANGE

        for x in range(0, day_range):
            d = base - datetime.timedelta(days=x)
            allowed_dates.append(d.strftime(date_format))
            dates.append({
                'value': d.strftime(date_format),
                'label': d.strftime(date_format_display)
            })

        if selected_date not in allowed_dates:
            messages.warning(self.request,
                             _('This dates is blocked you are not allowed to take attendance for this date.'))
            self.template_name = 'error.html'
            return {
            }

        return {
            'school_type': 'ALP',
            'attendance': attendance,
            'disable_attendance': disable_attendance,
            'current_level_section': current_level_section,
            'total': queryset.count(),
            'total_students': students.count() if students else 0,
            'students': students,
            'school': school,
            'level': level,
            'section': section,
            'dates': dates,
            'classrooms': EducationLevel.objects.all(),
            'sections': Section.objects.all(),
            'levels_by_sections': levels_by_sections,
            'selected_date': selected_date,
            'selected_date_view': selected_date_view,
        }


class AttendancesExportViewSet(mixins.ListModelMixin,
                               viewsets.GenericViewSet,
                               SuperuserRequiredMixin):
    """
    Provides API import attendance data
    """
    model = Attendance
    queryset = Attendance.objects.all().order_by('attendance_date')
    serializer_class = AttendanceExportSerializer
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        from_day = int(self.request.GET.get('from_day', 1))
        to_day = int(self.request.GET.get('to_day', 1))
        month = int(self.request.GET.get('month', 0))
        year = int(self.request.GET.get('year', 0))
        max_raw = int(self.request.GET.get('max', 500))
        queryset = self.queryset

        queryset = queryset.filter(attendance_date__month=month)
        queryset = queryset.filter(attendance_date__year=year)
        queryset = queryset.filter(attendance_date__day__gte=from_day, attendance_date__day__lte=to_day)

        if self.request.GET.get('offset', 0):
            offset = int(self.request.GET.get('offset', 0))
            limit = offset + max_raw
            return queryset[offset:limit]
        return []


class ExportView(LoginRequiredMixin, ListView):
    model = Attendance
    queryset = Attendance.objects.all()

    def get(self, request, *args, **kwargs):
        date_format = '%Y-%m-%d'
        current_date = datetime.datetime.now().strftime(date_format)
        selected_date = self.request.GET.get('date', current_date)
        school_type = self.request.GET.get('school_type', '2nd-shift')

        school = self.request.user.school_id
        data = export_attendance({'date': selected_date, 'school': school, 'school_type': school_type},
                                 return_data=True)
        # data = export_attendance({'from_date': '2017-10-01', 'to_date': '2017-12-15', 'school_type': '2nd-shift', 'gov': 4}, return_data=True)

        response = HttpResponse(
            data,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=attendance_' + selected_date + '.xlsx'
        return response



class MainAttendanceCreateView(LoginRequiredMixin, GroupRequiredMixin, CreateView, RequestConfig):
    form_class = MainAttendanceForm
    template_name = 'attendances/main_attendance_form.html'
    group_required = [u"CLM_ATTENDANCE"]

    def get_initial_student_formset(self, initial_records):
        attendance_student_inline_formset = inlineformset_factory(
            CLMAttendance,
            CLMAttendanceStudent,
            form=AttendanceStudentForm,
            extra=len(initial_records),
            fk_name='attendance_day',
            fields=('attended', 'absence_reason', 'absence_reason_other', 'student_id'),
            can_delete=False
        )
        return attendance_student_inline_formset(initial=initial_records)

    def get_student_formset(self, parameters):
        attendance_student_inline_formset = inlineformset_factory(
            CLMAttendance,
            CLMAttendanceStudent,
            form=AttendanceStudentForm,
            fk_name='attendance_day',
            fields=('attended', 'absence_reason', 'absence_reason_other', 'student_id'),
            can_delete=False
        )
        return attendance_student_inline_formset(parameters)

    def get_form_kwargs(self):
        kwargs = super(MainAttendanceCreateView, self).get_form_kwargs()
        if self.request.method == 'POST':
            kwargs['attendance_student_formset'] = self.get_student_formset(self.request.POST)
        else:
            queryset = Bridging.objects.none()
            school = self.request.GET.get('school', None)
            attendance_date = self.request.GET.get('attendance_date', None)
            if school is not None:
                school = int(school)
                registration_level = self.request.GET.get('registration_level', '')
                day_off = self.request.GET.get('day_off', '')
                if school > 0 and registration_level != '' and day_off == 'no':
                    queryset = Bridging.objects.filter(
                                                       round__current_round_bridging=True,
                                                       school=school,
                                                       registration_level=registration_level)
                    if attendance_date is not None:
                        queryset = queryset.filter(
                            Q(registration_date__isnull=True) | Q(registration_date__lte=attendance_date),
                                Q(dropout_date__isnull=True) | Q(dropout_date__gt=attendance_date)
                        )

                    queryset = queryset.order_by('student__first_name', 'student__father_name', 'student__last_name')
            data = []
            for line in queryset:
                student = {
                    'student_id': line.student.id,
                    'student_name': line.student.full_name,
                    'attended': 'yes'
                }
                data.append(student)
            kwargs['attendance_student_formset'] = self.get_initial_student_formset(data)

        if self.request.user.partner:
            kwargs['partner_id'] = self.request.user.partner.id
        school_id = 0
        if self.request.user.school:
            school_id = self.request.user.school.id
        kwargs['user_school_id'] = school_id

        kwargs['clm_bridging_all'] = self.request.user.groups.filter(name='CLM_BRIDGING_ALL').exists()

        round_id = 0
        current_round = CLMRound.objects.filter(current_round_bridging=True).first()
        if current_round:
            round_id = current_round.id
        kwargs['round_id'] = round_id
        return kwargs

    def post(self, request, *args, **kwargs):
        self.object = None
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        attendance_student_formset = self.get_student_formset(self.request.POST)
        student_count = len(attendance_student_formset)
        if self.request.POST['day_off'] == 'yes' and student_count > 0:
            form.add_error('day_off', 'Day is off')

        if form.is_valid() and attendance_student_formset.is_valid():
            return self.form_valid(form, attendance_student_formset)
        else:
            return self.form_invalid(form, attendance_student_formset)

    def get(self, request, *args, **kwargs):
        attendance_date = self.request.GET.get('attendance_date', '')
        school = self.request.GET.get('school', '')
        registration_level = self.request.GET.get('registration_level', '')

        attendance = None
        if school != '' and registration_level != '' and attendance_date != '':
            attendance = CLMAttendance.objects.filter(school=school,
                                                      registration_level=registration_level,
                                                      attendance_date=attendance_date,
                                                     ).values('id').first()
        if attendance:
            attendance_id = attendance['id']
            return redirect(reverse('attendances:main_attendance_edit', kwargs={'pk': attendance_id}))
        else:
            return super(MainAttendanceCreateView, self).get(request)

    def form_valid(self, form, attendance_student_formset):
        self.object = form.save(commit=False)
        self.object.save()
        # saving ProductMeta Instances
        for student_form in attendance_student_formset:
            student_form.instance.student_id = student_form.cleaned_data['student_id']
        attendance_students = attendance_student_formset.save(commit=False)

        current_round = CLMRound.objects.all()
        current_round = current_round.get(current_round_bridging=True)
        for attendance_student in attendance_students:
            attendance_student.attendance_day = self.object
            attendance_student.save()
            student_id = attendance_student.student_id
            registration = Bridging.objects.filter(student_id=student_id, round_id=current_round.id).first()
            update_student_consecutive_absences(student_id, current_round, registration)
            update_student_total_attandance(student_id, current_round, registration)

        messages.success(self.request, 'The attendance information was saved')
        return super(MainAttendanceCreateView, self).form_valid(form)

    def form_invalid(self, form, attendance_student_formset):
        return self.render_to_response(
            self.get_context_data(form=form,
                                  attendance_student_formset=attendance_student_formset
                                  )
        )

    def get_initial(self):
        initial_values = {}
        if self.request.GET.get('school', None):
            initial_values['school'] = int(self.request.GET.get('school', 0))
            if self.request.GET.get('registration_level', None):
                initial_values['registration_level'] = self.request.GET.get('registration_level', '')
            if self.request.GET.get('day_off', None):
                initial_values['day_off'] = self.request.GET.get('day_off', '')
            if self.request.GET.get('attendance_date', None):
                initial_values['attendance_date'] = self.request.GET.get('attendance_date', '')
        return initial_values

    def get_success_message(self):
        return "The attendance information was saved"

    def get_success_url(self):
        return reverse('attendances:main_attendance')


class MainAttendanceUpdateView(LoginRequiredMixin, UpdateView):
    model = CLMAttendance
    form_class = MainAttendanceForm
    template_name = 'attendances/main_attendance_form.html'
    group_required = [u"CLM_ATTENDANCE"]

    def get_success_url(self):
        return reverse('attendances:main_attendance')
        # return reverse('attendances:main_attendance_edit', args=[self.kwargs['pk']])

    def get_context_data(self, **kwargs):
        force_default_language(self.request)
        """Insert the form into the context dict."""
        if 'form' not in kwargs:
            kwargs['form'] = self.get_form()
        return super(MainAttendanceUpdateView, self).get_context_data(**kwargs)

    def get_form(self, form_class=None):
        attendance_id = self.kwargs['pk']
        instance = CLMAttendance.objects.get(id=attendance_id)
        partner_id = 0
        user_school_id = 0
        if self.request.user.partner:
            partner_id = self.request.user.partner.id
        if self.request.user.school:
            user_school_id = self.request.user.school.id

        round_id = 0
        current_round = CLMRound.objects.filter(current_round_bridging=True).first()
        if current_round:
            round_id = current_round.id

        clm_bridging_all = self.request.user.groups.filter(name='CLM_BRIDGING_ALL').exists()
        # messages.success(self.request, 'There is already an attendance record for this date.')
        if self.request.method == "POST":
            instance.save()
            form = MainAttendanceForm(self.request.POST, instance=instance,
                                      attendance_student_formset=self.get_student_formset(self.request.POST),
                                      partner_id=partner_id, user_school_id=user_school_id,round_id=round_id,
                                      clm_bridging_all=clm_bridging_all)
        else:
            form = MainAttendanceForm(instance=instance,
                                      attendance_student_formset=self.get_formset(attendance_id),
                                      partner_id=partner_id, user_school_id=user_school_id,round_id=round_id,
                                      clm_bridging_all=clm_bridging_all)

        form.helper.form_action = reverse('attendances:main_attendance_edit', args=[attendance_id])

        return form

    def post(self, request, *args, **kwargs):
        self.object = None
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        attendance_student_formset = self.get_student_formset(self.request.POST)
        student_count = len(attendance_student_formset)
        if self.request.POST['day_off'] == 'yes' and student_count > 0:
            form.add_error('day_off', 'Day is off')

        if form.is_valid() and attendance_student_formset.is_valid():
            return self.form_valid(form, attendance_student_formset)
        else:
            return self.form_invalid(form, attendance_student_formset)

    def form_valid(self, form, attendance_student_formset):
        self.object = form.save(commit=False)
        self.object.save()

        for student_form in attendance_student_formset:
            student_form.instance.id = student_form.cleaned_data['id'].id
            student_form.instance.student_id = student_form.cleaned_data['student_id']

        attendance_students = attendance_student_formset.save(commit=False)

        current_round = CLMRound.objects.all()
        current_round = current_round.get(current_round_bridging=True)
        for attendance_student in attendance_students:
            attendance_student.attendance_day = self.object
            attendance_student.save()
            student_id = attendance_student.student_id
            registration = Bridging.objects.filter(student_id=student_id, round_id=current_round.id).first()
            update_student_consecutive_absences(student_id, current_round, registration)
            update_student_total_attandance(student_id, current_round, registration)
        messages.success(self.request, 'The attendance information was saved')
        return super(MainAttendanceUpdateView, self).form_valid(form)

    def form_invalid(self, form, attendance_student_formset):
        return self.render_to_response(
            self.get_context_data(form=form,
                                  attendance_student_formset=attendance_student_formset
                                  )
        )

    def get_formset(self, attendance_id):
        queryset = CLMAttendanceStudent.objects.filter(attendance_day__id=attendance_id)
        queryset = queryset.order_by('student__first_name', 'student__father_name', 'student__last_name')

        data = []
        for line in queryset:
            student = {
                'id': line.id,
                'student_id': line.student.id,
                'student_name': line.student.full_name,
                'attended': line.attended,
                'absence_reason': line.absence_reason,
                'absence_reason_other': line.absence_reason_other
            }
            data.append(student)
        return self.get_initial_student_formset(data)

    def get_initial_student_formset(self, initial_records):
        attendance_student_inline_formset = inlineformset_factory(
            CLMAttendance,
            CLMAttendanceStudent,
            form=AttendanceStudentForm,
            extra=len(initial_records),
            fk_name='attendance_day',
            fields=('id', 'attended', 'absence_reason', 'absence_reason_other', 'student_id'),
            can_delete=False
        )
        return attendance_student_inline_formset(initial=initial_records)

    def get_student_formset(self, parameters):
        attendance_student_inline_formset = inlineformset_factory(
            CLMAttendance,
            CLMAttendanceStudent,
            form=AttendanceStudentForm,
            fk_name='attendance_day',
            fields=('id', 'attended', 'absence_reason', 'absence_reason_other', 'student_id'),
            can_delete=False
        )
        return attendance_student_inline_formset(parameters)


def update_student_consecutive_absences(student_id,current_round, registration):

    student_absences = CLMAttendanceStudent.objects.filter(
        attended='no',
        attendance_day__attendance_date__range=(current_round.start_date_bridging, current_round.end_date_bridging),
        student_id=student_id
    ).order_by('student__id', 'attendance_day__attendance_date').all()

    days_off = []
    if student_absences:
        school_id = student_absences[0].attendance_day.school.id
        registration_level = student_absences[0].attendance_day.registration_level
        days_off = CLMAttendance.objects.filter(day_off='yes', registration_level = registration_level, school= school_id ,
                                                    attendance_date__range=(current_round.start_date_bridging, current_round.end_date_bridging)
                                                ).order_by('attendance_date').values_list('attendance_date', flat=True)

        partner_id = registration.partner.id

        working_day_names = School.objects.filter(id=school_id).values_list('working_days', flat=True).first()

        current_student_absence = StudentConsecutiveAbsenceTracking()

        current_student_absence.initialise(student_id,partner_id,school_id,current_round,registration,working_day_names,days_off)

        for row in student_absences:

            absence_date = row.attendance_day.attendance_date
            current_student_absence.check_update_absence_date(absence_date)

        current_student_absence.check_add_absence_record()

        CLMStudentAbsences.objects.filter\
        (\
            student_id=student_id,round_id=current_round.id,\
            absence_starting_date__range = (current_round.start_date_bridging, current_round.end_date_bridging) \
        )\
        .exclude(absence_starting_date__in=current_student_absence.absence_dates)\
        .delete()


class StudentConsecutiveAbsenceTracking:

    def initialise(current_student_absence,student_id,partner_id,school_id,   round, registration, working_day_names, days_off):

        current_student_absence.absence_days = 0
        current_student_absence.absence_starting_date = None
        current_student_absence.absence_ending_date = None
        current_student_absence.current_absence_date = None

        current_student_absence.student_id = None
        current_student_absence.absence_dates = []
        current_student_absence.consecutive_dates = []

        current_student_absence.student_id = student_id
        current_student_absence.partner_id = partner_id
        current_student_absence.school_id = school_id
        current_student_absence.round = round
        current_student_absence.registration = registration

        current_student_absence.working_day_names = working_day_names
        current_student_absence.days_off = days_off

    def check_add_absence_record(current_student_absence):
        if current_student_absence.absence_starting_date is not None:
            current_student_absence.add_absence_record()
            current_student_absence.absence_dates.append(current_student_absence.absence_starting_date)

    def add_absence_record(current_student_absence):

        round_id = current_student_absence.round.id
        student_id = current_student_absence.student_id
        partner_id = current_student_absence.partner_id
        school_id = current_student_absence.school_id
        absence_starting_date = current_student_absence.absence_starting_date

        absence_exists = CLMStudentAbsences.objects.filter(student_id=student_id, round_id=round_id,
                                                           absence_starting_date=absence_starting_date).exists()
        registration = Bridging.objects.filter(student_id=student_id, round_id=round_id ).first()

        if absence_exists:
            absence = CLMStudentAbsences.objects.filter(student_id=student_id, round_id=round_id,
                                                        absence_starting_date=absence_starting_date).first()
        else:

            absence = CLMStudentAbsences(round_id=round_id,
                                         student_id=student_id,
                                         partner_id=partner_id,
                                         school_id=school_id,
                                         absence_starting_date=absence_starting_date,
                                         registration_id = registration.id,
                                         student_first_name=registration.student.first_name,
                                         student_father_name=registration.student.father_name,
                                         student_last_name=registration.student.last_name,
                                         school_name=registration.school.name,
                                         registation_level=registration.registration_level
                                         )


        start_date = current_student_absence.absence_starting_date
        end_date = current_student_absence.absence_ending_date
        current_student_absence.consecutive_dates = []
        delta = datetime.timedelta(days=1)
        while start_date <= end_date:
            if not current_student_absence.is_date_off_weekend(start_date, ):
                current_student_absence.consecutive_dates.append(str(start_date))
            start_date += delta

        absence.update_absence_statisics(current_student_absence.absence_days, current_student_absence.absence_ending_date, current_student_absence.consecutive_dates)
        absence.save()

    def check_update_absence_date(current_student_absence,absence_date):

        if current_student_absence.absence_starting_date is None:
            current_student_absence.absence_starting_date = absence_date
            current_student_absence.absence_ending_date = absence_date
            current_student_absence.absence_days = 1

        elif current_student_absence.all_range_off(current_student_absence.absence_ending_date,absence_date):

            current_student_absence.absence_ending_date = absence_date
            current_student_absence.absence_days += 1

        elif (absence_date != (current_student_absence.absence_ending_date + datetime.timedelta(days=1))):

            current_student_absence.check_add_absence_record()
            current_student_absence.absence_starting_date = absence_date
            current_student_absence.absence_ending_date = absence_date
            current_student_absence.absence_days = 1


    def all_range_off(current_student_absence,previous_absence_date,new_absence_date):

        result = True

        current_absence_date = previous_absence_date + datetime.timedelta(days=1)

        while(current_absence_date!=new_absence_date and result):

            result = result and current_student_absence.is_date_off_weekend(current_absence_date,)

            current_absence_date += datetime.timedelta(days=1)

        return result

    def is_date_off_weekend(current_student_absence,current_absence_date):

        day_name = current_absence_date.strftime("%A")
        working_day_names = current_student_absence.working_day_names
        days_off = current_student_absence.days_off

        return (day_name not in working_day_names) or (current_absence_date in days_off)


def update_student_total_attandance(student_id, current_round, registration):

    total_absence_days = CLMAttendanceStudent.objects.filter(
        attended='no',
        attendance_day__attendance_date__range=(current_round.start_date_bridging, current_round.end_date_bridging),
        student_id=student_id
    ).order_by('student__id', 'attendance_day__attendance_date').count()
    total_attendance_days = CLMAttendanceStudent.objects.filter(
        attended='yes',
        attendance_day__attendance_date__range=(current_round.start_date_bridging, current_round.end_date_bridging),
        student_id=student_id
    ).order_by('student__id', 'attendance_day__attendance_date').count()

    student_id = student_id
    round_id = current_round.id
    registration = registration

    absence_exists = CLMStudentTotalAttendance.objects.filter(student_id=student_id, registration_id=registration.id).exists()

    if absence_exists:
        absence = CLMStudentTotalAttendance.objects.filter(student_id=student_id, registration_id=registration.id).first()
    else:

        absence = CLMStudentTotalAttendance(round_id=round_id,
                                            student_id=student_id,
                                            partner_id=registration.partner.id,
                                            school_id=registration.school.id,
                                            registration_id=registration.id,
                                            student_first_name=registration.student.first_name,
                                            student_father_name=registration.student.father_name,
                                            student_last_name=registration.student.last_name,
                                            school_name=registration.school.name,
                                            registation_level=registration.registration_level
                                             )

    absence.update_absence_statisics(total_absence_days)
    absence.update_attendance_statisics(total_attendance_days)

    absence.save()


class AttendanceAbsenceView(FormView):
    template_name = 'attendances/attendance_Absence_form.html'
    form_class = AttendanceAbsenceForm
    success_url = '/attendances/attendance-absence/'

    def form_valid(self, form):
        return super(AttendanceAbsenceView, self).form_valid(form)


@login_required(login_url='/users/login')
def absence_export(request,number_of_absences, total_days):

    number_of_consecutive_absences = int(number_of_absences)
    number_of_total_absence = int(total_days)

    current_round = CLMRound.objects.all()
    current_round = current_round.get(current_round_bridging=True)
    round_id = current_round.id

    consecutive_student_id = CLMStudentAbsences.objects.filter(
        consecutive_absence_days__gte=number_of_consecutive_absences,
        round_id=round_id) \
        .values_list('student_id', flat=True)
    total_student_id = CLMStudentTotalAttendance.objects.filter(
        total_absence_days__gte=number_of_total_absence,
        round_id=round_id) \
        .values_list('student_id', flat=True)

    consecutive_student_list = list(consecutive_student_id)
    total_student_list = list(total_student_id)

    student_ids = consecutive_student_list + list(set(total_student_list) - set(consecutive_student_list))
    student_ids = list(set(student_ids))

    consecutive_absent_students = CLMStudentAbsences.objects.filter(student_id__in= student_ids)\
                                   .order_by('student_id','absence_starting_date').all()

    buffer = io.BytesIO()

    wb_student = xlwt.Workbook(encoding='utf-8', style_compression=2)
    ws = wb_student.add_sheet('Student')

    # Sheet header, first row
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns_titles = [
        'Student First Name',
        'Student Father Name',
        'Student Last Name',
        'School',
        'Registation Level',
        'Number of Consecutive Absences',
        'Consecutive Absences From',
        'Consecutive Absences To',
        'Absence Dates',
        'Total Attendance Days',
        'Total Absence Days',
    ]

    for col_num in range(len(columns_titles)):
        ws.write(0, col_num, columns_titles[col_num], font_style)

    row_num_student = 0
    font_style = xlwt.XFStyle()
    for row in consecutive_absent_students:
        row_num_student += 1
        ws.write(row_num_student, 0, row.student_first_name,font_style)
        ws.write(row_num_student, 1, row.student_father_name,font_style)
        ws.write(row_num_student, 2, row.student_last_name,font_style)
        ws.write(row_num_student, 3, row.school_name,font_style)
        ws.write(row_num_student, 4, row.registation_level,font_style)
        ws.write(row_num_student, 5, row.consecutive_absence_days,font_style)
        ws.write(row_num_student, 6, row.absence_starting_date,font_style)
        ws.write(row_num_student, 7, row.absence_ending_date,font_style)
        ws.write(row_num_student, 8, row.absence_dates,font_style)

        student_id= row.student_id
        total_absent_students = CLMStudentTotalAttendance.objects\
            .filter(student_id=student_id, round_id=round_id).last()

        ws.write(row_num_student, 9, total_absent_students.total_attendance_days,font_style)
        ws.write(row_num_student, 10, total_absent_students.total_absence_days,font_style)

    wb_student.save(buffer)

    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    response = FileResponse(buffer, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Absence.xls"'

    return response


from django.http import StreamingHttpResponse
from openpyxl import Workbook
from django.db import connection


def generate_workbook(headers, cursor, max_rows_per_sheet):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Attendance Records 1"
    worksheet.append(headers)

    sheet_count = 1
    row_count = 0
    chunk_size = 1000

    while True:
        rows = cursor.fetchmany(chunk_size)
        if not rows:
            break
        for row in rows:
            if row_count >= max_rows_per_sheet:
                sheet_count += 1
                sheet_title = "Attendance Records " + str(sheet_count)
                worksheet = workbook.create_sheet(title=sheet_title)
                worksheet.append(headers)
                row_count = 0
            worksheet.append(row)
            row_count += 1

    # Remove the default sheet if it exists
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    return workbook


def stream_workbook(workbook):
    from io import BytesIO
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer



def attendance_export_xlsx(request, **kwargs):
    month = kwargs.get('month')
    year = kwargs.get('year')

    current_round = CLMRound.objects.get(current_round_bridging=True)
    round_id = current_round.id

    cursor = connection.cursor()

    vw_data_str = "SELECT * FROM vw_bridging_attendance WHERE round_id = %s"
    if year:
        vw_data_str += " AND EXTRACT(YEAR FROM attendance_date) = " + year
    if month:
        vw_data_str += " AND EXTRACT(MONTH FROM attendance_date) = " + month

    query_params = [round_id]

    if not request.user.groups.filter(name='CLM_BRIDGING_ALL').exists():
        if request.user.partner_id:
            vw_data_str += " AND partner_id = %s"
            query_params.append(request.user.partner_id)
        if request.user.school_id:
            vw_data_str += " AND school_id = %s"
            query_params.append(request.user.school_id)

    vw_data_str += " ORDER BY attendance_date"
    cursor.execute(vw_data_str, query_params)

    headers = [col[0] for col in cursor.description]

    max_rows_per_sheet = 50000
    workbook = generate_workbook(headers, cursor, max_rows_per_sheet)
    buffer = stream_workbook(workbook)

    response = StreamingHttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=raw_attendance.xlsx'

    return response


@login_required(login_url='/users/login')
def attendance_export(request, **kwargs):
    try:
        user = request.user
        partner_name = user.partner.name if user.partner else ''

        month = kwargs.get('month')
        year = kwargs.get('year')

        cursor = connection.cursor()
        query_params = []
        vw_data_str = "SELECT * FROM vw_bridging_attendance WHERE round_id > 0"

        if year:
            vw_data_str += " AND EXTRACT(YEAR FROM attendance_date) = %s"
            query_params.append(year)
        if month:
            vw_data_str += " AND EXTRACT(MONTH FROM attendance_date) = %s"
            query_params.append(month)

        if not request.user.groups.filter(name='CLM_BRIDGING_ALL').exists():
            if request.user.partner_id:
                vw_data_str += " AND partner_id = %s"
                query_params.append(request.user.partner_id)
            if request.user.school_id:
                vw_data_str += " AND school_id = %s"
                query_params.append(request.user.school_id)

        vw_data_str += " ORDER BY attendance_date"

        cursor.execute(vw_data_str, query_params)
        att_data = cursor.fetchall()

        # Log the query for debugging purposes
        logging.debug("Executing query: %s", vw_data_str)
        logging.debug("Query params: %s", str(query_params))

        headers = [col[0] for col in cursor.description]

        # Create CSV
        csv_output = io.StringIO()
        csv_writer = csv.writer(csv_output)

        # Add BOM for Arabic text
        csv_output.write(codecs.BOM_UTF8.decode('utf-8'))
        csv_writer.writerow(headers)

        for row in att_data:
            encoded_row = []
            for cell in row:
                if isinstance(cell, (str, bytes)):  # Handle string and bytes
                    encoded_row.append(cell.decode('utf-8') if isinstance(cell, bytes) else cell)
                elif isinstance(cell, (datetime.date, datetime.datetime)):  # Convert date/datetime objects to string
                    encoded_row.append(cell.strftime('%Y-%m-%d'))
                else:  # Convert other data types to string
                    encoded_row.append(str(cell))

            csv_writer.writerow(encoded_row)

        unique_id = str(uuid.uuid4())
        file_name = "consecutive_absence_{}.csv".format(unique_id)
        file_path = os.path.join('export', file_name)

        # Save file
        default_storage.save(file_path, ContentFile(csv_output.getvalue().encode('utf-8')))

        # Store export history
        ExportHistory.objects.create(
            export_type='Bridging Absence Raw Data',
            created_by=user,
            partner_name=partner_name
        )

        return HttpResponse(file_name)


    except Exception as e:
        logging.error("An error occurred during the export process:")
        logging.error(traceback.format_exc())

        return HttpResponse("An error occurred: " + str(e), status=500)


@login_required(login_url='/users/login')
def total_attendance_export(request, **kwargs):
    try:
        user = request.user
        partner_name = user.partner.name if user.partner else ''

        round_id = kwargs.get('round')
        cursor = connection.cursor()
        query_params = [round_id]
        vw_data_str = "SELECT * FROM vw_bridging_attendance_total WHERE round_id = %s"

        if not request.user.groups.filter(name='CLM_BRIDGING_ALL').exists():
            if request.user.partner_id:
                vw_data_str += " AND partner_id = %s"
                query_params.append(request.user.partner_id)
            if request.user.school_id:
                vw_data_str += " AND school_id = %s"
                query_params.append(request.user.school_id)

        vw_data_str += " ORDER BY student_id"
        cursor.execute(vw_data_str, query_params)
        att_data = cursor.fetchall()

        logging.debug("Executing query: %s", vw_data_str)
        logging.debug("Query params: %s", str(query_params))

        headers = [col[0] for col in cursor.description]

        # Create CSV
        csv_output = io.StringIO()
        csv_writer = csv.writer(csv_output)

        # Add BOM for Arabic text
        csv_output.write(codecs.BOM_UTF8.decode('utf-8'))
        csv_writer.writerow(headers)  # Write headers

        for row in att_data:
            csv_writer.writerow(
                [cell.decode('utf-8') if isinstance(cell, bytes) else
                 cell.strftime('%Y-%m-%d') if isinstance(cell, (datetime.date, datetime.datetime)) else
                 str(cell) for cell in row]
            )

        unique_id = str(uuid.uuid4())
        file_name = "total_attendance_{}.csv".format(unique_id)
        file_path = os.path.join('export', file_name)

        # Save file
        default_storage.save(file_path, ContentFile(csv_output.getvalue().encode('utf-8')))

        # Store export history
        ExportHistory.objects.create(
            export_type='Bridging Attendance Total',
            created_by=user,
            partner_name=partner_name
        )

        return HttpResponse(file_name)

    except Exception as e:
        logging.error("An error occurred during the export process:")
        logging.error(traceback.format_exc())
        return HttpResponse("An error occurred: " + str(e), status=500)


@login_required(login_url='/users/login')
def consecutive_absence_export(request, **kwargs):
    try:
        user = request.user
        partner_name = user.partner.name if user.partner else ''

        round_id = kwargs.get('round')
        cursor = connection.cursor()
        query_params = [round_id]
        vw_data_str = "SELECT * FROM vw_bridging_absence_consecutive WHERE round_id = %s"

        if not request.user.groups.filter(name='CLM_BRIDGING_ALL').exists():
            if request.user.partner_id:
                vw_data_str += " AND partner_id = %s"
                query_params.append(request.user.partner_id)
            if request.user.school_id:
                vw_data_str += " AND school_id = %s"
                query_params.append(request.user.school_id)

        vw_data_str += " ORDER BY student_id"
        cursor.execute(vw_data_str, query_params)
        att_data = cursor.fetchall()

        logging.debug("Executing query: %s", vw_data_str)
        logging.debug("Query params: %s", str(query_params))

        headers = [col[0] for col in cursor.description]

        # Create CSV
        csv_output = io.StringIO()
        csv_writer = csv.writer(csv_output)

        # Add BOM for Arabic text
        csv_output.write(codecs.BOM_UTF8.decode('utf-8'))
        csv_writer.writerow(headers)  # Write headers

        for row in att_data:
            csv_writer.writerow(
                [cell.decode('utf-8') if isinstance(cell, bytes) else
                 cell.strftime('%Y-%m-%d') if isinstance(cell, (datetime.date, datetime.datetime)) else
                 str(cell) for cell in row]
            )

        unique_id = str(uuid.uuid4())
        file_name = "consecutive_absence_{}.csv".format(unique_id)
        file_path = os.path.join('export', file_name)

        # Save file
        default_storage.save(file_path, ContentFile(csv_output.getvalue().encode('utf-8')))

        # Store export history
        ExportHistory.objects.create(
            export_type='Bridging Absence Consecutive',
            created_by=user,
            partner_name=partner_name
        )

        return HttpResponse(file_name)

    except Exception as e:
        logging.error("An error occurred during the export process:")
        logging.error(traceback.format_exc())
        return HttpResponse("An error occurred: " + str(e), status=500)

@login_required(login_url='/users/login')
def mscc_attendance_export(request, **kwargs):
    try:
        user = request.user
        partner_name = user.partner.name if user.partner else ''

        month = kwargs.get('month')
        year = kwargs.get('year')
        # round = kwargs.get('round')
        education_program = kwargs.get('education_program', '')
        class_section = kwargs.get('class_section', '')
        partner_id = kwargs.get('partner', '')
        center_id = kwargs.get('center', '')

        if education_program == 'none':
            education_program = ''
        if class_section == 'none':
            class_section = ''
        if partner_id == 'none':
            partner_id = ''
        if center_id == 'none':
            center_id = ''

        cursor = connection.cursor()
        query_params = []
        vw_data_str = "SELECT * FROM vw_mscc_attendance_data WHERE attendance_id > 0"

        if year:
            vw_data_str += " AND EXTRACT(YEAR FROM attendance_date) = %s"
            query_params.append(year)
        if month:
            vw_data_str += " AND EXTRACT(MONTH FROM attendance_date) = %s"
            query_params.append(month)

        if education_program:
            vw_data_str += " AND education_program = %s"
            query_params.append(education_program)

        if class_section:
            vw_data_str += " AND class_section = %s"
            query_params.append(class_section)

        if partner_id:
            vw_data_str += " AND partner_id = %s"
            query_params.append(partner_id)

        if center_id:
            vw_data_str += " AND center_id = %s"
            query_params.append(center_id)

        if not request.user.groups.filter(name='MSCC_UNICEF').exists():
            if request.user.partner_id:
                vw_data_str += " AND partner_id = %s"
                query_params.append(request.user.partner_id)

            if request.user.center_id:
                vw_data_str += " AND center_id = %s"
                query_params.append(request.user.center_id)

        vw_data_str += " ORDER BY attendance_date"

        cursor.execute(vw_data_str, query_params)
        att_data = cursor.fetchall()

        logging.debug("Executing query: %s", vw_data_str)
        logging.debug("Query params: %s", str(query_params))

        headers = [col[0] for col in cursor.description]

        # Create CSV
        csv_output = io.StringIO()
        csv_writer = csv.writer(csv_output)

        # Add BOM for Arabic text
        csv_output.write(codecs.BOM_UTF8.decode('utf-8'))
        csv_writer.writerow(headers)  # Write headers

        for row in att_data:
            encoded_row = []
            for cell in row:
                if isinstance(cell, (str, bytes)):  # Handle string and bytes
                    encoded_row.append(cell.decode('utf-8') if isinstance(cell, bytes) else cell)
                elif isinstance(cell, (datetime.date, datetime.datetime)):  # Convert date/datetime objects to string
                    encoded_row.append(cell.strftime('%Y-%m-%d'))
                else:  # Convert other data types to string
                    encoded_row.append(str(cell))
            csv_writer.writerow(encoded_row)

        unique_id = str(uuid.uuid4())
        file_name = "mscc_raw_attendance_data_{}.csv".format(unique_id)
        file_path = os.path.join('export', file_name)

        # Save file
        default_storage.save(file_path, ContentFile(csv_output.getvalue().encode('utf-8')))

        # Store export history
        ExportHistory.objects.create(
            export_type='Makani Raw Attendance',
            created_by=user,
            partner_name=partner_name
        )

        return HttpResponse(file_name)

    except Exception as e:
        logging.error("An error occurred during the export process:")
        logging.error(traceback.format_exc())
        return HttpResponse("An error occurred: " + str(e), status=500)


@login_required(login_url='/users/login')
def mscc_total_attendance_export(request, **kwargs):
    try:
        user = request.user
        partner_name = user.partner.name if user.partner else ''

        month = kwargs.get('month')
        year = kwargs.get('year')
        # round = kwargs.get('round')
        education_program = kwargs.get('education_program', '')
        class_section = kwargs.get('class_section', '')
        partner_id = kwargs.get('partner', '')
        center_id = kwargs.get('center', '')

        if education_program == 'none':
            education_program = ''
        if class_section == 'none':
            class_section = ''
        if partner_id == 'none':
            partner_id = ''
        if center_id == 'none':
            center_id = ''

        cursor = connection.cursor()
        query_params = []
        vw_data_str = "SELECT * FROM vw_mscc_attendance_absence WHERE center_id > 0"

        if year:
            vw_data_str += " AND attendance_year = %s"
            query_params.append(year)
        if month:
            vw_data_str += " AND attendance_month = %s"
            query_params.append(month)

        if education_program:
            vw_data_str += " AND education_program = %s"
            query_params.append(education_program)

        if class_section:
            vw_data_str += " AND class_section = %s"
            query_params.append(class_section)

        if partner_id:
            vw_data_str += " AND partner_id = %s"
            query_params.append(partner_id)

        if center_id:
            vw_data_str += " AND center_id = %s"
            query_params.append(center_id)

        if not request.user.groups.filter(name='MSCC_UNICEF').exists():
            if request.user.partner_id:
                vw_data_str += " AND partner_id = %s"
                query_params.append(request.user.partner_id)

            if request.user.center_id:
                vw_data_str += " AND center_id = %s"
                query_params.append(request.user.center_id)

        vw_data_str += " ORDER BY child_id"
        cursor.execute(vw_data_str, query_params)
        att_data = cursor.fetchall()

        logging.debug("Executing query: %s", vw_data_str)
        logging.debug("Query params: %s", str(query_params))

        headers = [col[0] for col in cursor.description]

        # Create CSV
        csv_output = io.StringIO()
        csv_writer = csv.writer(csv_output)

        # Add BOM for Arabic text
        csv_output.write(codecs.BOM_UTF8.decode('utf-8'))
        csv_writer.writerow(headers)  # Write headers

        for row in att_data:
            encoded_row = []
            for cell in row:
                if isinstance(cell, (str, bytes)):  # Handle string and bytes
                    encoded_row.append(cell.decode('utf-8') if isinstance(cell, bytes) else cell)
                elif isinstance(cell,
                                (datetime.date, datetime.datetime)):  # Convert date/datetime objects to string
                    encoded_row.append(cell.strftime('%Y-%m-%d'))
                else:  # Convert other data types to string
                    encoded_row.append(str(cell))
            csv_writer.writerow(encoded_row)

        unique_id = str(uuid.uuid4())
        file_name = "mscc_total_attendance_data_{}.csv".format(unique_id)
        file_path = os.path.join('export', file_name)

        # Save file
        default_storage.save(file_path, ContentFile(csv_output.getvalue().encode('utf-8')))

        # Store export history
        ExportHistory.objects.create(
            export_type='Makani Total Attendance',
            created_by=user,
            partner_name=partner_name
        )

        return HttpResponse(file_name)

    except Exception as e:
        logging.error("An error occurred during the export process:")
        logging.error(traceback.format_exc())
        return HttpResponse("An error occurred: " + str(e), status=500)

@login_required(login_url='/users/login')
def mscc_total_attendance_export1(request, **kwargs):
    try:
        user = request.user
        partner_name = ''
        if user.partner:
            partner_name = user.partner.name

        month = kwargs.get('month')
        year = kwargs.get('year')
        # round = kwargs.get('round')
        education_program = kwargs.get('education_program', '')
        class_section = kwargs.get('class_section', '')
        partner_id = kwargs.get('partner', '')
        center_id = kwargs.get('center', '')

        if education_program == 'none':
            education_program = ''
        if class_section == 'none':
            class_section = ''
        if partner_id == 'none':
            partner_id = ''
        if center_id == 'none':
            center_id = ''

        cursor = connection.cursor()
        query_params = []
        vw_data_str = "SELECT * FROM vw_mscc_attendance_absence WHERE center_id > 0"

        if year:
            vw_data_str += " AND attendance_year = %s"
            query_params.append(year)
        if month:
            vw_data_str += " AND attendance_month = %s"
            query_params.append(month)

        if education_program:
            vw_data_str += " AND education_program = %s"
            query_params.append(education_program)

        if class_section:
            vw_data_str += " AND class_section = %s"
            query_params.append(class_section)

        if partner_id:
            vw_data_str += " AND partner_id = %s"
            query_params.append(partner_id)

        if center_id:
            vw_data_str += " AND center_id = %s"
            query_params.append(center_id)

        if not request.user.groups.filter(name='MSCC_UNICEF').exists():
            if request.user.center_id:
                vw_data_str += " AND center_id = %s"
                query_params.append(request.user.center_id)

        vw_data_str += " ORDER BY child_id"
        cursor.execute(vw_data_str, query_params)
        att_data = cursor.fetchall()

        # Log the query for debugging purposes
        logging.debug("Executing query: %s", vw_data_str)
        logging.debug("Query params: %s", str(query_params))

        headers = [col[0] for col in cursor.description]
        zip_output = io.BytesIO()
        with zipfile.ZipFile(zip_output, 'w') as zf:
            # Create CSV for vw_mscc_data
            csv_mscc_output = io.StringIO()
            csv_writer = csv.writer(csv_mscc_output)

            # Add BOM to handle Arabic text correctly
            csv_mscc_output.write(codecs.BOM_UTF8.decode('utf-8'))
            csv_writer.writerow(headers)  # Write headers

            for row in att_data:
                encoded_row = []
                for cell in row:
                    if isinstance(cell, (str, bytes)):  # Handle string and bytes
                        encoded_row.append(cell.decode('utf-8') if isinstance(cell, bytes) else cell)
                    elif isinstance(cell,
                                    (datetime.date, datetime.datetime)):  # Convert date/datetime objects to string
                        encoded_row.append(cell.strftime('%Y-%m-%d'))
                    else:  # Convert other data types to string
                        encoded_row.append(str(cell))
                csv_writer.writerow(encoded_row)

            # Add CSV to ZIP
            zf.writestr('mscc_total_attendance_data.csv', csv_mscc_output.getvalue())

        unique_id = str(uuid.uuid4())
        file_name = "out_file_{}.zip".format(unique_id)
        file_path = os.path.join('export', file_name)

        default_storage.save(file_path, ContentFile(zip_output.getvalue()))
        ExportHistory.objects.create(
            export_type='Makani Total Attendance',
            created_by=user,
            partner_name=partner_name
        )
        return HttpResponse(file_name)

    except Exception as e:
        logging.error("An error occurred during the export process:")
        logging.error(traceback.format_exc())

        return HttpResponse("An error occurred: " + str(e), status=500)
